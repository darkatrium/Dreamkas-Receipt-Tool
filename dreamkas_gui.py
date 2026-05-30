#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Simple GUI launcher for Dreamkas Receipt Tool.

This GUI does not replace the fiscalization logic. It launches the proven
console script in a separate terminal, opens Excel, and opens work folders.
"""

from __future__ import annotations

import os
import subprocess
import sys
import shutil
from pathlib import Path
import tkinter as tk
from tkinter import filedialog, messagebox


APP_DIR = Path(__file__).resolve().parent
SCRIPT = APP_DIR / "dreamkas_receipt.py"
DEFAULT_EXCEL = APP_DIR / "dreamkas_receipt_template.xlsx"


def create_default_excel_template(excel_path: Path) -> None:
    """Create a basic Excel template if it is missing."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    except Exception as exc:
        messagebox.showerror("Dreamkas", f"openpyxl is required to create Excel template:\n{exc}")
        return

    excel_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Чек"

    rows = [
        ("Тип оплаты", "Безнал"),
        ("Email покупателя", ""),
        ("Телефон покупателя", ""),
        ("Тип покупателя", "Физлицо"),
        ("Наименование юрлица / ИП", ""),
        ("ИНН юрлица / ИП", ""),
        ("Система налогообложения", "SIMPLE_WO"),
        ("Имя кассира", ""),
        ("Наименование", "Тип (Услуга = 1 | Товар = 0)", "Количество", "Цена", "Ставка НДС (0 - Без НДС)"),
        ("", "", "", "", ""),
    ]

    for r_idx, row in enumerate(rows, start=1):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx).value = value

    for col, width in {"A": 34, "B": 28, "C": 14, "D": 14, "E": 24}.items():
        ws.column_dimensions[col].width = width

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    meta_fill = PatternFill("solid", fgColor="F2F2F2")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in range(1, 9):
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=1).fill = meta_fill
        ws.cell(row=row, column=1).border = border
        ws.cell(row=row, column=2).border = border

    for col in range(1, 6):
        cell = ws.cell(row=9, column=col)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    for row in range(10, 80):
        for col in range(1, 6):
            ws.cell(row=row, column=col).border = border

    ws.freeze_panes = "A10"
    wb.save(excel_path)


def ensure_default_excel_template() -> Path:
    """Ensure default Excel template exists next to the GUI/script/EXE."""
    if DEFAULT_EXCEL.exists():
        return DEFAULT_EXCEL

    # If running under PyInstaller, try bundled resource first.
    meipass = getattr(sys, "_MEIPASS", None)
    if meipass:
        bundled = Path(meipass) / "dreamkas_receipt_template.xlsx"
        if bundled.exists():
            try:
                shutil.copy2(bundled, DEFAULT_EXCEL)
                return DEFAULT_EXCEL
            except Exception:
                pass

    create_default_excel_template(DEFAULT_EXCEL)
    return DEFAULT_EXCEL




def find_python() -> str:
    candidates = [
        r"C:\Python314\python.exe",
        sys.executable,
        "py",
        "python",
    ]
    for cmd in candidates:
        try:
            if Path(cmd).exists():
                return cmd
            subprocess.run([cmd, "--version"], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL, check=False)
            return cmd
        except Exception:
            continue
    return "python"


def open_path(path: Path) -> None:
    path = path.resolve()
    if not path.exists():
        messagebox.showwarning("Dreamkas", f"Path not found:\n{path}")
        return
    if os.name == "nt":
        os.startfile(str(path))  # type: ignore[attr-defined]
    elif sys.platform == "darwin":
        subprocess.Popen(["open", str(path)])
    else:
        subprocess.Popen(["xdg-open", str(path)])


class App(tk.Tk):
    def __init__(self) -> None:
        super().__init__()
        self.title("Dreamkas Receipt Tool")
        self.geometry("720x360")
        self.resizable(False, False)

        self.excel_var = tk.StringVar(value=str(ensure_default_excel_template()))
        self.status_var = tk.StringVar(value="Ready")

        self._build_ui()

    def _build_ui(self) -> None:
        pad = {"padx": 10, "pady": 6}

        title = tk.Label(self, text="Dreamkas Receipt Tool v6.33", font=("Segoe UI", 16, "bold"))
        title.pack(anchor="w", padx=14, pady=(14, 4))

        frame = tk.Frame(self)
        frame.pack(fill="x", **pad)

        tk.Label(frame, text="Excel file:").grid(row=0, column=0, sticky="w")
        entry = tk.Entry(frame, textvariable=self.excel_var, width=76)
        entry.grid(row=1, column=0, sticky="we", padx=(0, 8))
        tk.Button(frame, text="Choose...", command=self.choose_excel).grid(row=1, column=1)

        btns = tk.Frame(self)
        btns.pack(fill="x", **pad)

        tk.Button(btns, text="Open Excel", width=18, command=self.open_excel).grid(row=0, column=0, padx=4, pady=4)
        tk.Button(btns, text="Start console tool", width=18, command=self.start_console).grid(row=0, column=1, padx=4, pady=4)
        tk.Button(btns, text="Open receipts", width=18, command=lambda: open_path(APP_DIR / "receipts_pdf")).grid(row=0, column=2, padx=4, pady=4)
        tk.Button(btns, text="Open DB folder", width=18, command=lambda: open_path(APP_DIR / "db")).grid(row=1, column=0, padx=4, pady=4)
        tk.Button(btns, text="Open logs", width=18, command=lambda: open_path(APP_DIR / "logs")).grid(row=1, column=1, padx=4, pady=4)
        tk.Button(btns, text="Open settings", width=18, command=lambda: open_path(APP_DIR / "settings.txt")).grid(row=1, column=2, padx=4, pady=4)

        info = (
            "The console window remains the main safe workflow: shop/device selection, "
            "SALE/REFUND mode, precheck confirmation, SQLite journal, PDF/TXT/QR output."
        )
        tk.Label(self, text=info, wraplength=680, justify="left").pack(anchor="w", padx=14, pady=(10, 4))

        status = tk.Label(self, textvariable=self.status_var, anchor="w", relief="sunken")
        status.pack(side="bottom", fill="x")

    def choose_excel(self) -> None:
        path = filedialog.askopenfilename(
            title="Choose Excel receipt template",
            initialdir=str(APP_DIR),
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if path:
            self.excel_var.set(path)

    def open_excel(self) -> None:
        path = Path(self.excel_var.get())
        if path.name == "dreamkas_receipt_template.xlsx" and not path.exists():
            path = ensure_default_excel_template()
            self.excel_var.set(str(path))
        open_path(path)

    def start_console(self) -> None:
        if not SCRIPT.exists():
            messagebox.showerror("Dreamkas", "dreamkas_receipt.py not found")
            return
        excel = Path(self.excel_var.get())
        py = find_python()

        if os.name == "nt":
            cmd = f'cd /d "{APP_DIR}" && "{py}" "dreamkas_receipt.py" --excel "{excel}" && pause'
            subprocess.Popen(["cmd.exe", "/k", cmd])
        else:
            subprocess.Popen([py, str(SCRIPT), "--excel", str(excel)], cwd=str(APP_DIR))

        self.status_var.set("Console tool started")


if __name__ == "__main__":
    App().mainloop()
