#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Dreamkas receipt sender from Excel template.

Что делает:
1) читает TXT-настройки формата TOKEN_KEY = VALUE;
   при первом запуске сам спрашивает DREAMKAS_TOKEN и сохраняет settings.txt;
   при каждом запуске загружает список магазинов и касс из Dreamkas API,
   предлагает выбрать магазин/кассу и сохраняет SHOP_ID / DEVICE_ID в settings.txt;
2) читает Excel-файл с параметрами чека и позициями;
   поддерживает новый шаблон: отдельная строка Email и отдельная строка Телефон;
   старый шаблон с одной строкой Email/Телефон тоже поддерживается;
3) показывает предчек в терминале;
4) после подтверждения отправляет чек в Dreamkas API;
5) сохраняет локальный журнал в SQLite;
6) опрашивает Dreamkas до SUCCESS или ERROR;
7) при SUCCESS сохраняет TXT-чек и QR-код ФНС;
8) умеет делать полный или частичный возврат ранее сохраненного чека из SQLite.

Запуск:
    python dreamkas_receipt.py --settings settings.txt --excel dreamkas_receipt_template.xlsx

Зависимости:
    pip install requests openpyxl qrcode[pil]
"""

from __future__ import annotations

import argparse
import base64
import getpass
import hashlib
import hmac
import ssl
import json
import logging
import os
import re
import shutil
import sqlite3
import sys
import time
import traceback
import uuid
import smtplib
from email.message import EmailMessage
from copy import deepcopy
from dataclasses import dataclass
from datetime import datetime, timedelta
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Any, Optional

try:
    import requests
    from openpyxl import Workbook, load_workbook
    import qrcode
except ImportError as exc:
    print("Не хватает Python-библиотек.")
    print("Установи зависимости командой:")
    print("    pip install requests openpyxl qrcode[pil]")
    print(f"\nТехническая ошибка импорта: {exc}")
    input("\nНажмите Enter для выхода...")
    sys.exit(1)


API_DEFAULT_BASE_URL = "https://kabinet.dreamkas.ru/api"
APP_VERSION = "6.32"
DB_FILE = "dreamkas_receipts.sqlite3"

DB_DIR = "db"
LOG_DIR = "logs"
RECEIPTS_TXT_DIR = "receipts_txt"
RECEIPTS_QR_DIR = "receipts_qr"
RECEIPTS_PDF_DIR = "receipts_pdf"
TEMPLATES_DIR = "templates"


@dataclass
class ReceiptPosition:
    name: str
    item_type: str
    quantity: Decimal
    price_kopecks: int
    price_sum_kopecks: int
    tax: str


@dataclass
class ReceiptDraft:
    external_id: str
    payment_type: str
    buyer_email: Optional[str]
    buyer_phone: Optional[str]
    buyer_type: str
    buyer_legal_name: Optional[str]
    buyer_inn: Optional[str]
    tax_mode: str
    cashier_name: str
    positions: list[ReceiptPosition]
    total_kopecks: int
    payload: dict[str, Any]
    parent_external_id: Optional[str] = None
    parent_receipt_id: Optional[str] = None
    refund_mode: Optional[str] = None
    refunded_positions: Optional[list[int]] = None

@dataclass
class ReceiptInputMeta:
    """Реквизиты чека, которые вводятся в терминале."""
    payment_type: str
    buyer_email: Optional[str]
    buyer_phone: Optional[str]
    buyer_type: str
    buyer_legal_name: Optional[str]
    buyer_inn: Optional[str]
    tax_mode: str
    cashier_name: str


@dataclass
class SbpPaymentResult:
    """Результат платежной сессии СБП T-Банка."""
    payment_id: str
    order_id: str
    status: str
    qr_payload: Optional[str]
    qr_path: Optional[str]
    init_response: dict[str, Any]
    qr_response: dict[str, Any]
    state_response: dict[str, Any]



class DreamkasError(Exception):
    pass


class OperatorCancelled(DreamkasError):
    """
    Оператор нажал ESC во время ожидания платежа или фискализации.

    Это локальная отмена ожидания. Если задание уже отправлено в Dreamkas,
    оно может продолжить выполняться на стороне кассы.
    """
    pass


CANCEL_INPUTS = {"0", "q", "й", "quit", "exit", "esc", "отмена", "назад", "выход"}


def is_cancel_input(raw: str) -> bool:
    return str(raw or "").strip().lower() in CANCEL_INPUTS


def raise_input_cancelled(stage: str = "ввод реквизитов") -> None:
    raise OperatorCancelled(f"Оператор отменил операцию: {stage}")


def input_or_cancel(prompt: str, *, allow_zero_as_value: bool = False, stage: str = "ввод реквизитов") -> str:
    """
    input() с поддержкой отмены.

    По умолчанию 0/q/отмена/назад/выход отменяют операцию.
    Для меню, где 0 является рабочим значением, используйте allow_zero_as_value=True.
    """
    raw = input(prompt)
    if is_cancel_input(raw) and not (allow_zero_as_value and raw.strip() == "0"):
        raise_input_cancelled(stage)
    return raw


def prompt_continue_or_cancel(message: str) -> None:
    raw = input(message)
    if is_cancel_input(raw):
        raise_input_cancelled("ожидание заполнения Excel")


def wait_key(message: str = "Нажмите любую клавишу для выхода...") -> None:
    """Кроссплатформенная пауза перед закрытием окна."""
    print(message)
    try:
        if os.name == "nt":
            import msvcrt  # type: ignore
            msvcrt.getch()
        else:
            input()
    except Exception:
        pass


def operator_cancel_requested() -> bool:
    """
    Проверяет, нажал ли оператор ESC.

    На Windows используется msvcrt без блокировки.
    На других ОС возвращает False, чтобы не ломать выполнение.
    """
    try:
        import msvcrt  # type: ignore
        if not msvcrt.kbhit():
            return False
        key = msvcrt.getch()
        return key == b"\x1b"
    except Exception:
        return False


def raise_if_operator_cancelled(stage: str) -> None:
    if operator_cancel_requested():
        raise OperatorCancelled(f"Оператор отменил ожидание: {stage}")


def interruptible_sleep(seconds: int, stage: str) -> None:
    """
    Сон с проверкой ESC, чтобы оператор мог вернуться в главное меню
    во время ожидания платежа или ответа Dreamkas.
    """
    seconds = max(1, int(seconds))
    deadline = time.time() + seconds
    while time.time() < deadline:
        raise_if_operator_cancelled(stage)
        time.sleep(min(0.2, max(0.0, deadline - time.time())))
    raise_if_operator_cancelled(stage)


def print_cancel_hint() -> None:
    print("Нажмите ESC для отмены ожидания и возврата в главное меню.")


def get_app_dir() -> Path:
    """
    Возвращает папку приложения.

    Для обычного .py это папка скрипта.
    Для PyInstaller EXE это папка, где лежит .exe.
    """
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent


def get_resource_path(filename: str) -> Optional[Path]:
    """
    Ищет файл-ресурс внутри PyInstaller _MEIPASS или рядом со скриптом.

    При сборке EXE Excel-шаблон добавляется через:
        --add-data "dreamkas_receipt_template.xlsx;."
    """
    meipass = getattr(sys, "_MEIPASS", None)
    if meipass:
        candidate = Path(meipass) / filename
        if candidate.exists():
            return candidate

    candidate = get_app_dir() / filename
    if candidate.exists():
        return candidate

    return None


def resolve_app_file(path: Path) -> Path:
    """
    Делает относительный путь относительным к папке приложения, а не к случайной cwd.

    Это важно для EXE: если программу открыть двойным кликом, рабочая папка
    может отличаться от папки, где лежит .exe.
    """
    path = path.expanduser()
    if path.is_absolute():
        return path
    return get_app_dir() / path


def create_default_excel_template(excel_path: Path) -> None:
    """
    Создает Excel-шаблон только для товарных позиций.

    Все реквизиты чека вводятся в терминале:
    - тип оплаты;
    - email/телефон покупателя;
    - тип покупателя;
    - ИНН/название юрлица;
    - СНО;
    - кассир.

    В шаблоне есть выпадающие списки:
    - колонка B: 1 = услуга, 0 = товар;
    - колонка E: ставка НДС.
    """
    excel_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Позиции"

    headers = [
        "Наименование",
        "Тип (Услуга = 1 | Товар = 0)",
        "Количество",
        "Цена",
        "Ставка НДС (0 - Без НДС)",
    ]

    for col, value in enumerate(headers, start=1):
        ws.cell(row=1, column=col).value = value

    for row in range(2, 80):
        for col in range(1, 6):
            ws.cell(row=row, column=col).value = None

    widths = {"A": 42, "B": 28, "C": 14, "D": 14, "E": 24}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.worksheet.datavalidation import DataValidation

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col in range(1, 6):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    for row in range(2, 80):
        for col in range(1, 6):
            ws.cell(row=row, column=col).border = border

    item_type_validation = DataValidation(
        type="list",
        formula1='"1,0"',
        allow_blank=False,
        showErrorMessage=True,
        errorTitle="Неверный тип позиции",
        error="Выберите 1 для услуги или 0 для товара.",
    )
    item_type_validation.promptTitle = "Тип позиции"
    item_type_validation.prompt = "1 = услуга, 0 = товар"
    ws.add_data_validation(item_type_validation)
    item_type_validation.add("B2:B79")

    vat_validation = DataValidation(
        type="list",
        formula1='"0,5,7,10,22,5/105,7/107,10/110,22/122"',
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Неверная ставка НДС",
        error="Выберите ставку НДС из списка.",
    )
    vat_validation.promptTitle = "Ставка НДС"
    vat_validation.prompt = "0 = без НДС; также доступны 5, 7, 10, 22 и расчетные ставки."
    ws.add_data_validation(vat_validation)
    vat_validation.add("E2:E79")

    ws.freeze_panes = "A2"
    wb.save(excel_path)

def ensure_excel_template_available(excel_path: Path) -> Path:
    """
    Проверяет наличие Excel-шаблона рядом с программой.

    Если шаблона нет:
    1. пробует восстановить его из встроенного ресурса PyInstaller;
    2. если ресурса нет — создает новый шаблон программно.
    """
    excel_path = resolve_app_file(excel_path)

    if excel_path.exists():
        return excel_path

    print("\nExcel-шаблон не найден рядом с программой.")
    print(f"Ожидаемый путь: {excel_path}")

    resource = get_resource_path("dreamkas_receipt_template.xlsx")
    if resource and resource.exists() and resource.resolve() != excel_path.resolve():
        try:
            excel_path.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(resource, excel_path)
            print("Excel-шаблон восстановлен из встроенного ресурса EXE.")
            print(f"Создан файл: {excel_path}")
            return excel_path
        except Exception as exc:
            print(f"Не удалось скопировать встроенный шаблон: {exc}")

    create_default_excel_template(excel_path)
    print("Создан новый Excel-шаблон по умолчанию.")
    print(f"Создан файл: {excel_path}")
    return excel_path



def open_excel_for_user_fill(excel_path: Path) -> None:
    """
    Открывает Excel-шаблон перед формированием предчека.

    Сценарий нужен для режима фискализации: пользователь сначала заполняет
    шаблон, сохраняет его, затем возвращается в терминал и нажимает клавишу.
    После этого скрипт читает файл и показывает предчек.
    """
    excel_path = ensure_excel_template_available(excel_path)

    print("\nОткрываю Excel-шаблон для заполнения...")
    print(f"Файл: {excel_path}")

    try:
        if os.name == "nt":
            os.startfile(str(excel_path))  # type: ignore[attr-defined]
        elif sys.platform == "darwin":
            import subprocess
            subprocess.Popen(["open", str(excel_path)])
        else:
            import subprocess
            subprocess.Popen(["xdg-open", str(excel_path)])
        print("Excel-файл открыт.")
    except Exception as exc:
        print("\nНе удалось автоматически открыть Excel-файл.")
        print("Откройте его вручную, заполните и сохраните:")
        print(f"    {excel_path}")
        print(f"Техническая причина: {exc}")

    print("\nЗаполните Excel-шаблон и сохраните его.")
    print("После сохранения вернитесь в это окно терминала.")
    prompt_continue_or_cancel("После того как все будет сделано, нажмите Enter для вывода предчека или 0 для отмены: ")



SETTINGS_ORDER = [
    "DREAMKAS_TOKEN",
    "SHOP_ID",
    "SHOP_NAME",
    "DEVICE_ID",
    "DEVICE_NAME",
    "DEFAULT_CASHIER_NAME",
    "DEFAULT_PAYMENT_TYPE",
    "CASHLESS_PAYMENT_PROVIDER",
    "DEFAULT_TAX_MODE",
    "API_BASE_URL",
    "OPERATION_TYPE",
    "TIMEOUT_MINUTES",
    "POLL_INTERVAL_SECONDS",
    "REQUEST_TIMEOUT_SECONDS",
    "DUPLICATE_WINDOW_MINUTES",
    "DUPLICATE_LOOKBACK_DAYS",
    "CLEAR_EXCEL_AFTER_SUCCESS",
    "PDF_ENABLED",
    "EMAIL_ENABLED_AFTER_SUCCESS",
    "SMTP_HOST",
    "SMTP_PORT",
    "SMTP_USER",
    "SMTP_PASSWORD_ENC",
    "SMTP_FROM",
    "SMTP_SSL",
    "SMTP_STARTTLS",
    "SMTP_VERIFY_CERT",
    "SEND_LEGAL_ENTITY_TAGS_TO_DREAMKAS",
    "LEGAL_ENTITY_LOOKUP_ENABLED",
    "LEGAL_ENTITY_LOOKUP_PROVIDER",
    "DADATA_TOKEN",
    "DADATA_BRANCH_TYPE",
    "DADATA_UPDATE_EXCEL_NAME",
    "TBANK_SBP_ENABLED",
    "TBANK_TERMINAL_KEY",
    "TBANK_PASSWORD_ENC",
    "TBANK_PAY_TYPE",
    "TBANK_PAYMENT_TIMEOUT_MINUTES",
    "TBANK_POLL_INTERVAL_SECONDS",
    "TBANK_REQUEST_TIMEOUT_SECONDS",
    "TBANK_QR_DATA_TYPE",
    "TBANK_DESCRIPTION_PREFIX",
    "TBANK_B2B_SBP_ENABLED",
    "TBANK_B2B_API_TOKEN_ENC",
    "TBANK_B2B_ACCOUNT_NUMBER",
    "TBANK_B2B_TTL_DAYS",
    "TBANK_B2B_VAT",
    "TBANK_B2B_REDIRECT_URL",
    "TBANK_B2B_PURPOSE_PREFIX",
]

DEFAULT_SETTINGS = {
    "DREAMKAS_TOKEN": "",
    "SHOP_ID": "",
    "SHOP_NAME": "",
    "DEVICE_ID": "",
    "DEVICE_NAME": "",
    "DEFAULT_CASHIER_NAME": "",
    "DEFAULT_PAYMENT_TYPE": "CASHLESS",
    "CASHLESS_PAYMENT_PROVIDER": "EXTERNAL_TERMINAL",
    "DEFAULT_TAX_MODE": "SIMPLE_WO",
    "API_BASE_URL": API_DEFAULT_BASE_URL,
    "OPERATION_TYPE": "SALE",
    "TIMEOUT_MINUTES": "5",
    "POLL_INTERVAL_SECONDS": "20",
    "REQUEST_TIMEOUT_SECONDS": "30",
    "DUPLICATE_WINDOW_MINUTES": "10",
    "DUPLICATE_LOOKBACK_DAYS": "30",
    "CLEAR_EXCEL_AFTER_SUCCESS": "1",
    "PDF_ENABLED": "1",
    "EMAIL_ENABLED_AFTER_SUCCESS": "0",
    "SMTP_HOST": "",
    "SMTP_PORT": "587",
    "SMTP_USER": "",
    "SMTP_PASSWORD_ENC": "",
    "SMTP_FROM": "",
    "SMTP_SSL": "0",
    "SMTP_STARTTLS": "1",
    "SMTP_VERIFY_CERT": "1",
    "SEND_LEGAL_ENTITY_TAGS_TO_DREAMKAS": "1",
    "LEGAL_ENTITY_LOOKUP_ENABLED": "1",
    "LEGAL_ENTITY_LOOKUP_PROVIDER": "DADATA",
    "DADATA_TOKEN": "",
    "DADATA_BRANCH_TYPE": "MAIN",
    "DADATA_UPDATE_EXCEL_NAME": "1",
    "TBANK_SBP_ENABLED": "0",
    "TBANK_TERMINAL_KEY": "",
    "TBANK_PASSWORD_ENC": "",
    "TBANK_PAY_TYPE": "O",
    "TBANK_PAYMENT_TIMEOUT_MINUTES": "10",
    "TBANK_POLL_INTERVAL_SECONDS": "5",
    "TBANK_REQUEST_TIMEOUT_SECONDS": "30",
    "TBANK_QR_DATA_TYPE": "PAYLOAD",
    "TBANK_DESCRIPTION_PREFIX": "Оплата заказа",
    "TBANK_B2B_SBP_ENABLED": "0",
    "TBANK_B2B_API_TOKEN_ENC": "",
    "TBANK_B2B_ACCOUNT_NUMBER": "",
    "TBANK_B2B_TTL_DAYS": "1",
    "TBANK_B2B_VAT": "0",
    "TBANK_B2B_REDIRECT_URL": "",
    "TBANK_B2B_PURPOSE_PREFIX": "Оплата по счету",
}

TAX_MODE_CHOICES: list[tuple[str, str, str]] = [
    ("DEFAULT", "По умолчанию / общая система на кассе", "DEFAULT; для ОСН в Dreamkas API используется именно DEFAULT"),
    ("SIMPLE", "Упрощённая система — доходы", "УСН доходы"),
    ("SIMPLE_WO", "Упрощённая система — доходы минус расходы", "УСН доходы-расходы"),
    ("AGRICULT", "Единый сельскохозяйственный налог", "ЕСХН"),
    ("PATENT", "Патентная система налогообложения", "ПСН / патент"),
]


def tax_mode_label(value: str) -> str:
    value = str(value or "").strip().upper()
    for code, title, short in TAX_MODE_CHOICES:
        if value == code:
            return f"{title} ({code})"
    return value or "-"



def normalize_tax_mode_for_dreamkas(value: str) -> str:
    """
    Финальная защита перед отправкой в Dreamkas API.

    Допустимые taxMode для /api/receipts:
    DEFAULT, SIMPLE, SIMPLE_WO, AGRICULT, PATENT.
    """
    value = str(value or "").strip().upper()
    if value == "OSN":
        return "DEFAULT"
    if value in {"DEFAULT", "SIMPLE", "SIMPLE_WO", "AGRICULT", "PATENT"}:
        return value
    return "DEFAULT"

def print_inline_status(message: str) -> None:
    """
    Перезаписывает текущую строку статуса без переноса строки.
    Используется при ожидании оплаты СБП, чтобы QR-код не сдвигался вниз.
    """
    try:
        width = shutil.get_terminal_size((120, 20)).columns
    except Exception:
        width = 120
    clean = str(message).replace("\n", " ")
    limit = max(width - 1, 20)
    if len(clean) > limit:
        clean = clean[: max(limit - 3, 20)] + "..."
    print("\r" + clean.ljust(limit), end="", flush=True)


def finish_inline_status() -> None:
    """Завершает строку inline-статуса нормальным переносом."""
    print("", flush=True)



def bool_setting(settings: dict[str, str], key: str, default: bool = False) -> bool:
    raw = str(settings.get(key, "")).strip().lower()
    if not raw:
        return default
    return raw in {"1", "true", "yes", "y", "да", "д", "on", "вкл"}


def ensure_work_folders(workdir: Path) -> dict[str, Path]:
    """Создает рабочие папки и возвращает их пути."""
    paths = {
        "root": workdir,
        "db": workdir / DB_DIR,
        "logs": workdir / LOG_DIR,
        "receipts_txt": workdir / RECEIPTS_TXT_DIR,
        "receipts_qr": workdir / RECEIPTS_QR_DIR,
        "receipts_pdf": workdir / RECEIPTS_PDF_DIR,
        "templates": workdir / TEMPLATES_DIR,
    }
    for path in paths.values():
        if path != workdir:
            path.mkdir(parents=True, exist_ok=True)

    # Мягкая миграция старой базы из корня в db/, чтобы история чеков не пропала.
    old_db = workdir / DB_FILE
    new_db = paths["db"] / DB_FILE
    if old_db.exists() and not new_db.exists():
        shutil.copy2(old_db, new_db)
    return paths


def configure_logging(log_dir: Path) -> None:
    log_dir.mkdir(parents=True, exist_ok=True)
    log_path = log_dir / f"dreamkas_{datetime.now().strftime('%Y-%m-%d')}.log"
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s %(levelname)s %(message)s",
        handlers=[
            logging.FileHandler(log_path, encoding="utf-8"),
        ],
        force=True,
    )
    logging.info("Dreamkas Receipt Tool v%s started", APP_VERSION)


def log_json(title: str, data: Any) -> None:
    try:
        logging.info("%s\n%s", title, json.dumps(data, ensure_ascii=False, indent=2, default=str))
    except Exception:
        logging.info("%s: %r", title, data)




def load_settings(path: Path) -> dict[str, str]:
    """
    Читает настройки формата KEY = VALUE.

    Если файл отсутствует, возвращает пустой словарь. Это нужно для первого запуска:
    скрипт сможет сам спросить токен и создать settings.txt.
    """
    if not path.exists():
        return {}

    settings: dict[str, str] = {}
    for line_no, raw in enumerate(path.read_text(encoding="utf-8-sig").splitlines(), start=1):
        line = raw.strip()
        if not line or line.startswith("#") or line.startswith(";"):
            continue
        if "=" not in line:
            raise ValueError(f"Ошибка в settings.txt, строка {line_no}: нет символа '='")
        key, value = line.split("=", 1)
        key = key.strip().upper()
        value = value.strip().strip('"').strip("'")
        if not key:
            raise ValueError(f"Ошибка в settings.txt, строка {line_no}: пустой ключ")
        settings[key] = value
    return settings


def save_settings(path: Path, settings: dict[str, str]) -> None:
    """Сохраняет настройки в понятном TXT-формате KEY = VALUE."""
    merged = dict(DEFAULT_SETTINGS)
    merged.update({k.upper(): str(v).strip() for k, v in settings.items()})

    lines: list[str] = [
        "# Настройки Dreamkas receipt sender",
        "# Формат: TOKEN_KEY = VALUE",
        "#",
        "# DREAMKAS_TOKEN создается автоматически при первом запуске,",
        "# если он отсутствует в этом файле.",
        "# SHOP_ID и DEVICE_ID выбираются заново при каждом запуске",
        "# из списка магазинов и касс Dreamkas.",
        "",
    ]

    for key in SETTINGS_ORDER:
        lines.append(f"{key} = {merged.get(key, '')}")

    extra_keys = sorted(k for k in merged.keys() if k not in SETTINGS_ORDER)
    if extra_keys:
        lines.append("")
        lines.append("# Дополнительные параметры")
        for key in extra_keys:
            lines.append(f"{key} = {merged[key]}")

    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def ask_token_from_user() -> str:
    """Запрашивает токен у пользователя. getpass скрывает ввод, input — запасной вариант."""
    print("\nПервый запуск или в settings.txt нет DREAMKAS_TOKEN.")
    print("Введите API-токен Dreamkas. Он будет сохранен в settings.txt.")
    try:
        token = getpass.getpass("DREAMKAS_TOKEN: ").strip()
    except Exception:
        token = input("DREAMKAS_TOKEN: ").strip()

    if not token:
        raise ValueError("Токен Dreamkas не введен")
    return token


def ensure_token(settings_path: Path, settings: dict[str, str]) -> dict[str, str]:
    """
    Проверяет наличие DREAMKAS_TOKEN.
    Если токена нет — спрашивает его и сохраняет settings.txt.
    """
    token = settings.get("DREAMKAS_TOKEN", "").strip()
    if token:
        return settings

    settings = dict(settings)
    settings["DREAMKAS_TOKEN"] = ask_token_from_user()
    save_settings(settings_path, settings)
    print(f"\nТокен сохранен в файл настроек: {settings_path}")

    return settings




def smtp_password_kdf(token: str, salt: bytes) -> bytes:
    """
    Делает ключ шифрования SMTP-пароля из DREAMKAS_TOKEN.

    Важно: это переносимая схема. Если кто-то получит settings.txt целиком,
    где лежит и DREAMKAS_TOKEN, и SMTP_PASSWORD_ENC, он сможет расшифровать пароль.
    Защита рассчитана на случайный просмотр/копирование пароля, а не на кражу всего файла.
    """
    if not token:
        raise ValueError("Нельзя расшифровать SMTP-пароль: нет DREAMKAS_TOKEN")
    return hashlib.pbkdf2_hmac("sha256", token.encode("utf-8"), salt, 200_000, dklen=32)


def xor_bytes(data: bytes, stream: bytes) -> bytes:
    return bytes(a ^ b for a, b in zip(data, stream))


def hmac_stream(key: bytes, nonce: bytes, length: int) -> bytes:
    """Генерирует псевдослучайный поток байтов на HMAC-SHA256."""
    output = bytearray()
    counter = 0
    while len(output) < length:
        block = hmac.new(key, nonce + counter.to_bytes(4, "big"), hashlib.sha256).digest()
        output.extend(block)
        counter += 1
    return bytes(output[:length])


def encrypt_smtp_password(plain_password: str, dreamkas_token: str) -> str:
    """Шифрует SMTP-пароль ключом, производным от DREAMKAS_TOKEN."""
    salt = os.urandom(16)
    nonce = os.urandom(16)
    key = smtp_password_kdf(dreamkas_token, salt)
    plain = plain_password.encode("utf-8")
    cipher = xor_bytes(plain, hmac_stream(key, nonce, len(plain)))
    payload_without_tag = b"SMTPPW1" + salt + nonce + cipher
    tag = hmac.new(key, payload_without_tag, hashlib.sha256).digest()
    payload = salt + nonce + tag + cipher
    return "enc-v1:" + base64.urlsafe_b64encode(payload).decode("ascii")


def decrypt_smtp_password(encrypted_value: str, dreamkas_token: str) -> str:
    """Расшифровывает SMTP_PASSWORD_ENC. При смене токена старый пароль уже не открыть."""
    value = (encrypted_value or "").strip()
    if not value:
        return ""
    if not value.startswith("enc-v1:"):
        # Для совместимости: если вдруг передали старое открытое значение.
        return value
    raw = base64.urlsafe_b64decode(value.split(":", 1)[1].encode("ascii"))
    if len(raw) < 16 + 16 + 32:
        raise ValueError("Некорректное значение SMTP_PASSWORD_ENC")
    salt = raw[:16]
    nonce = raw[16:32]
    tag = raw[32:64]
    cipher = raw[64:]
    key = smtp_password_kdf(dreamkas_token, salt)
    payload_without_tag = b"SMTPPW1" + salt + nonce + cipher
    expected_tag = hmac.new(key, payload_without_tag, hashlib.sha256).digest()
    if not hmac.compare_digest(tag, expected_tag):
        raise ValueError("SMTP_PASSWORD_ENC не расшифрован: токен изменился или данные повреждены")
    plain = xor_bytes(cipher, hmac_stream(key, nonce, len(cipher)))
    return plain.decode("utf-8")


def get_smtp_password(settings_path: Path, settings: dict[str, str], *, ask_if_missing: bool = True) -> str:
    """
    Возвращает SMTP-пароль.

    Поддерживает мягкую миграцию:
    - если есть старый SMTP_PASSWORD в открытом виде — шифрует его в SMTP_PASSWORD_ENC;
    - затем удаляет SMTP_PASSWORD из settings.txt.
    """
    token = settings.get("DREAMKAS_TOKEN", "").strip()
    encrypted = settings.get("SMTP_PASSWORD_ENC", "").strip()
    if encrypted:
        return decrypt_smtp_password(encrypted, token)

    plain_legacy = settings.get("SMTP_PASSWORD", "").strip()
    if plain_legacy:
        settings["SMTP_PASSWORD_ENC"] = encrypt_smtp_password(plain_legacy, token)
        settings.pop("SMTP_PASSWORD", None)
        save_settings(settings_path, settings)
        print("SMTP_PASSWORD найден в открытом виде, зашифрован и перенесен в SMTP_PASSWORD_ENC.")
        return plain_legacy

    if not ask_if_missing:
        return ""

    if bool_setting(settings, "EMAIL_ENABLED_AFTER_SUCCESS", False):
        print("\nSMTP-пароль не найден. Введите его один раз — он будет сохранен в settings.txt в зашифрованном виде.")
        password = getpass.getpass("SMTP_PASSWORD: ").strip()
        if password:
            settings["SMTP_PASSWORD_ENC"] = encrypt_smtp_password(password, token)
            settings.pop("SMTP_PASSWORD", None)
            save_settings(settings_path, settings)
            print("SMTP-пароль зашифрован и сохранен как SMTP_PASSWORD_ENC.")
            return password
    return ""


def clear_smtp_settings(settings_path: Path, settings: dict[str, str], reason: str = "") -> dict[str, str]:
    """Удаляет SMTP-настройки из settings.txt, если токен Dreamkas недействителен."""
    updated = dict(settings)
    for key in list(updated.keys()):
        if key.startswith("SMTP_"):
            updated.pop(key, None)
    updated["EMAIL_ENABLED_AFTER_SUCCESS"] = "0"
    save_settings(settings_path, updated)
    logging.warning("SMTP settings cleared. Reason: %s", reason)
    print("\nВНИМАНИЕ: токен Dreamkas не прошел проверку.")
    print("SMTP-настройки удалены из settings.txt, отправка email отключена.")
    if reason:
        print(f"Причина: {reason}")
    return updated


def is_token_auth_error(exc: Exception) -> bool:
    text = str(exc).lower()
    return "http 401" in text or "http 403" in text or "unauthorized" in text or "forbidden" in text


def smtp_ssl_context(settings: dict[str, str]) -> ssl.SSLContext:
    verify = bool_setting(settings, "SMTP_VERIFY_CERT", True)
    if verify:
        return ssl.create_default_context()
    print("Внимание: SMTP_VERIFY_CERT = 0, проверка SSL-сертификата SMTP отключена.")
    return ssl._create_unverified_context()

def get_required(settings: dict[str, str], key: str) -> str:
    value = settings.get(key.upper(), "").strip()
    if not value:
        raise ValueError(f"В файле настроек отсутствует обязательный параметр {key}")
    return value


def get_int(settings: dict[str, str], key: str, default: int) -> int:
    value = settings.get(key.upper(), "").strip()
    if not value:
        return default
    try:
        return int(value)
    except ValueError as exc:
        raise ValueError(f"Параметр {key} должен быть числом") from exc


def read_meta_cell(ws: Any, row: int) -> str:
    """
    Поддерживает два варианта шаблона:
    1) A1 = 'Безнал'
    2) A1 = 'Тип оплаты:', B1 = 'Безнал'
    """
    a = ws.cell(row=row, column=1).value
    b = ws.cell(row=row, column=2).value
    if b is not None and str(b).strip() != "":
        return str(b).strip()
    if a is not None:
        return str(a).strip()
    return ""


def decimal_from_cell(value: Any, field_name: str, row: int) -> Decimal:
    if value is None or str(value).strip() == "":
        raise ValueError(f"Строка {row}: поле '{field_name}' пустое")
    if isinstance(value, Decimal):
        return value
    text = str(value).strip()
    text = text.replace("\u00a0", " ").replace(" ", "")
    text = text.replace("₽", "").replace("руб.", "").replace("руб", "")
    text = text.replace(",", ".")
    try:
        return Decimal(text)
    except InvalidOperation as exc:
        raise ValueError(f"Строка {row}: поле '{field_name}' не является числом: {value}") from exc


def rub_to_kopecks(value: Any, row: int) -> int:
    rub = decimal_from_cell(value, "Цена", row)
    if rub < 0:
        raise ValueError(f"Строка {row}: цена не может быть отрицательной")
    return int((rub * Decimal("100")).quantize(Decimal("1"), rounding=ROUND_HALF_UP))


def normalize_quantity(value: Any, row: int) -> Decimal:
    qty = decimal_from_cell(value, "Количество", row)
    if qty <= 0:
        raise ValueError(f"Строка {row}: количество должно быть больше 0")
    return qty


def decimal_to_json_number(value: Decimal) -> int | float:
    if value == value.to_integral_value():
        return int(value)
    return float(value)


def parse_payment_type(raw: str) -> str:
    text = raw.strip().lower().replace("ё", "е").replace(" ", "")
    mapping = {
        "cash": "CASH",
        "нал": "CASH",
        "наличные": "CASH",
        "наличными": "CASH",
        "cashless": "CASHLESS",
        "безнал": "CASHLESS",
        "безналичные": "CASHLESS",
        "безналичными": "CASHLESS",
        "карта": "CASHLESS",
        "картой": "CASHLESS",
        "sbp": "SBP",
        "сбп": "SBP",
        "qr": "SBP",
        "куар": "SBP",
        "tbank": "SBP",
        "тбанк": "SBP",
        "т-банк": "SBP",
    }
    if raw.strip().upper() in {"CASH", "CASHLESS", "SBP", "PREPAID", "CREDIT", "CONSIDERATION"}:
        return raw.strip().upper()
    if text in mapping:
        return mapping[text]
    raise ValueError("Тип оплаты должен быть 'Наличные', 'Безнал' или 'СБП'")


def normalize_cashless_payment_provider(raw: str) -> str:
    """
    Нормализует способ приема безналичной оплаты.

    EXTERNAL_TERMINAL — внешний банковский терминал / эквайринг.
    TBANK_SBP — интеграция СБП через T-Банк.
    """
    text = str(raw or "").strip().upper().replace(" ", "_").replace("-", "_")
    mapping = {
        "1": "EXTERNAL_TERMINAL",
        "EXTERNAL": "EXTERNAL_TERMINAL",
        "EXTERNAL_TERMINAL": "EXTERNAL_TERMINAL",
        "TERMINAL": "EXTERNAL_TERMINAL",
        "CARD_TERMINAL": "EXTERNAL_TERMINAL",
        "ACQUIRING": "EXTERNAL_TERMINAL",
        "ВНЕШНИЙ": "EXTERNAL_TERMINAL",
        "ВНЕШНИЙ_ТЕРМИНАЛ": "EXTERNAL_TERMINAL",
        "ТЕРМИНАЛ": "EXTERNAL_TERMINAL",
        "ЭКВАЙРИНГ": "EXTERNAL_TERMINAL",
        "2": "TBANK_SBP",
        "TBANK": "TBANK_SBP",
        "TBANK_SBP": "TBANK_SBP",
        "T_BANK": "TBANK_SBP",
        "TINKOFF": "TBANK_SBP",
        "SBP": "TBANK_SBP",
        "СБП": "TBANK_SBP",
        "ТБАНК": "TBANK_SBP",
        "Т_БАНК": "TBANK_SBP",
        "Т_БАНК_СБП": "TBANK_SBP",
    }
    return mapping.get(text, "EXTERNAL_TERMINAL")


def cashless_provider_label(provider: str) -> str:
    provider = normalize_cashless_payment_provider(provider)
    if provider == "TBANK_SBP":
        return "T-Банк СБП"
    return "внешний терминал"


def payment_label(payment_type: str, settings: Optional[dict[str, str]] = None) -> str:
    payment_type = str(payment_type or "").upper()
    if payment_type == "CASH":
        return "Наличные"
    if payment_type == "SBP":
        return "Безнал — T-Банк СБП"
    if payment_type == "CASHLESS":
        if settings:
            return f"Безнал — {cashless_provider_label(settings.get('CASHLESS_PAYMENT_PROVIDER', 'EXTERNAL_TERMINAL'))}"
        return "Безнал"
    return payment_type

def parse_tax_mode(raw: str) -> str:
    """
    Нормализует систему налогообложения для Dreamkas API.

    Dreamkas /api/receipts принимает только:
    DEFAULT, SIMPLE, SIMPLE_WO, AGRICULT, PATENT.

    Важное исправление:
    ОСН / общая система отправляется как DEFAULT, а не как OSN.
    """
    original = str(raw or "").strip()
    text_norm = original.lower().replace("ё", "е").replace(" ", "").replace("-", "").replace("_", "")

    number_mapping = {
        "0": "DEFAULT",
        "1": "SIMPLE",
        "2": "SIMPLE_WO",
        "3": "AGRICULT",
        "4": "PATENT",
    }
    if text_norm in number_mapping:
        return number_mapping[text_norm]

    direct = original.upper().strip()
    if direct == "OSN":
        return "DEFAULT"
    if direct == "ENVD":
        raise ValueError("ЕНВД не поддерживается Dreamkas API для /api/receipts. Выберите DEFAULT или другую доступную СНО.")
    if direct in {"DEFAULT", "SIMPLE", "SIMPLE_WO", "AGRICULT", "PATENT"}:
        return direct

    mapping = {
        "поумолчанию": "DEFAULT",
        "default": "DEFAULT",
        "касса": "DEFAULT",
        "какнакассе": "DEFAULT",

        "осн": "DEFAULT",
        "общая": "DEFAULT",
        "общаясистема": "DEFAULT",
        "общаясистеманалогообложения": "DEFAULT",
        "osn": "DEFAULT",

        "усн": "SIMPLE",
        "упрощенная": "SIMPLE",
        "упрощенка": "SIMPLE",
        "доходы": "SIMPLE",
        "усндоходы": "SIMPLE",
        "simple": "SIMPLE",

        "усндоходыминусрасходы": "SIMPLE_WO",
        "доходыминусрасходы": "SIMPLE_WO",
        "упрощеннаядоходыминусрасходы": "SIMPLE_WO",
        "упрощенкадоходыминусрасходы": "SIMPLE_WO",
        "simplewo": "SIMPLE_WO",
        "simplewithout": "SIMPLE_WO",

        "есхн": "AGRICULT",
        "сельхоз": "AGRICULT",
        "сельскохозяйственный": "AGRICULT",
        "единыйсельскохозяйственныйналог": "AGRICULT",
        "agricult": "AGRICULT",

        "патент": "PATENT",
        "псн": "PATENT",
        "патентная": "PATENT",
        "патентнаясистема": "PATENT",
        "patent": "PATENT",
    }

    if text_norm in {"енвд", "вмененка", "вмененный", "envd"}:
        raise ValueError("ЕНВД не поддерживается Dreamkas API для /api/receipts. Выберите DEFAULT или другую доступную СНО.")

    if text_norm in mapping:
        return mapping[text_norm]

    raise ValueError("Неизвестная система налогообложения. Выберите вариант из списка цифрой.")


def parse_item_type(raw: Any, row: int) -> str:
    text = str(raw).strip().lower().replace(" ", "")
    if text in {"1", "услуга", "service"}:
        return "SERVICE"
    if text in {"0", "товар", "countable", "шт", "штучный"}:
        return "COUNTABLE"
    raise ValueError(f"Строка {row}: тип должен быть 1=услуга или 0=товар")


def parse_tax(raw: Any, row: int) -> str:
    if raw is None or str(raw).strip() == "":
        return "NDS_NO_TAX"
    text = str(raw).strip().upper().replace(" ", "").replace("%", "")
    text = text.replace(",", ".")
    aliases = {
        "0": "NDS_NO_TAX",
        "БЕЗНДС": "NDS_NO_TAX",
        "NO": "NDS_NO_TAX",
        "NDS_NO_TAX": "NDS_NO_TAX",
        "5": "NDS_5",
        "7": "NDS_7",
        "10": "NDS_10",
        "22": "NDS_22",
        "105": "NDS_5_CALCULATED",
        "5/105": "NDS_5_CALCULATED",
        "107": "NDS_7_CALCULATED",
        "7/107": "NDS_7_CALCULATED",
        "110": "NDS_10_CALCULATED",
        "10/110": "NDS_10_CALCULATED",
        "122": "NDS_22_CALCULATED",
        "22/122": "NDS_22_CALCULATED",
        "NDS_0": "NDS_0",
        "NDS_5": "NDS_5",
        "NDS_7": "NDS_7",
        "NDS_10": "NDS_10",
        "NDS_22": "NDS_22",
        "NDS_5_CALCULATED": "NDS_5_CALCULATED",
        "NDS_7_CALCULATED": "NDS_7_CALCULATED",
        "NDS_10_CALCULATED": "NDS_10_CALCULATED",
        "NDS_22_CALCULATED": "NDS_22_CALCULATED",
    }
    if text in aliases:
        return aliases[text]
    raise ValueError(f"Строка {row}: неизвестная ставка НДС: {raw}")


def split_buyer_contact_optional(raw: str) -> tuple[Optional[str], Optional[str]]:
    """
    Разбирает строку, где может быть email, телефон или оба значения.
    Возвращает (email, phone). Если строка пустая — возвращает (None, None).
    """
    email: Optional[str] = None
    phone: Optional[str] = None
    raw = (raw or "").strip()
    if not raw:
        return None, None

    parts = re.split(r"[;,/\s]+", raw)
    for part in parts:
        part = part.strip()
        if not part:
            continue
        if "@" in part:
            email = part
        else:
            digits = re.sub(r"\D", "", part)
            if digits.startswith("8") and len(digits) == 11:
                digits = "7" + digits[1:]
            if digits.startswith("7") and len(digits) == 11:
                phone = "+" + digits
            elif part.startswith("+"):
                phone = part
    return email, phone


def split_buyer_contact(raw: str) -> tuple[Optional[str], Optional[str]]:
    email, phone = split_buyer_contact_optional(raw)
    if not email and not phone:
        raise ValueError("Нужно указать email покупателя или телефон в формате +7XXXXXXXXXX")
    return email, phone


def looks_like_header_row(ws: Any, row: int) -> bool:
    """Проверяет, похожа ли строка на шапку таблицы товарных позиций."""
    values = [str(ws.cell(row=row, column=col).value or "").strip().lower() for col in range(1, 6)]
    joined = " ".join(values).replace("ё", "е")
    return (
        "наименование" in joined
        and "тип" in joined
        and "кол" in joined
        and "цен" in joined
        and "ндс" in joined
    )


def read_buyer_contacts_new_template(ws: Any) -> tuple[Optional[str], Optional[str]]:
    """
    Новый шаблон:
      2 строка — Email покупателя
      3 строка — Телефон покупателя

    Для защиты от ошибок разрешаем ввести телефон в строку email или email в строку телефона:
    скрипт всё равно попробует распознать значения.
    """
    email_1, phone_1 = split_buyer_contact_optional(read_meta_cell(ws, 2))
    email_2, phone_2 = split_buyer_contact_optional(read_meta_cell(ws, 3))

    buyer_email = email_1 or email_2
    buyer_phone = phone_1 or phone_2

    if not buyer_email and not buyer_phone:
        raise ValueError("В новом Excel-шаблоне нужно заполнить Email покупателя во 2-й строке или Телефон покупателя в 3-й строке")

    return buyer_email, buyer_phone



def normalize_buyer_inn(raw: Any) -> Optional[str]:
    """Возвращает ИНН покупателя только цифрами или None."""
    text = str(raw or "").strip()
    if not text:
        return None
    digits = re.sub(r"\D", "", text)
    return digits or None


def normalize_buyer_type(raw: Any, legal_name: Optional[str] = None, buyer_inn: Optional[str] = None) -> str:
    """
    Определяет тип покупателя.

    Поддерживаемые значения в Excel:
    - Физлицо / ФЛ / физ / individual / person
    - Юрлицо / ЮЛ / организация / компания / ИП / legal / company / b2b

    Если тип не указан, но заполнены название или ИНН юрлица — считаем покупателя юрлицом.
    """
    text = str(raw or "").strip().lower().replace("ё", "е")
    if legal_name or buyer_inn:
        return "LEGAL_ENTITY"
    if not text:
        return "INDIVIDUAL"
    legal_markers = ["юр", "юл", "орган", "компан", "ооо", "ип", "legal", "company", "b2b"]
    individual_markers = ["физ", "фл", "част", "individual", "person"]
    if any(marker in text for marker in legal_markers):
        return "LEGAL_ENTITY"
    if any(marker in text for marker in individual_markers):
        return "INDIVIDUAL"
    # Неизвестное значение оставляем как физлицо, но валидация даст предупреждение.
    return "INDIVIDUAL"


def dadata_headers(token: str) -> dict[str, str]:
    return {
        "Content-Type": "application/json",
        "Accept": "application/json",
        "Authorization": f"Token {token}",
    }


def extract_dadata_party_name(suggestion: dict[str, Any]) -> Optional[str]:
    """
    Достаёт человекочитаемое название организации/ИП из ответа DaData.
    """
    data = suggestion.get("data") if isinstance(suggestion, dict) else None
    if not isinstance(data, dict):
        return str(suggestion.get("value") or "").strip() or None

    name = data.get("name")
    if isinstance(name, dict):
        for key in ("short_with_opf", "full_with_opf", "short", "full"):
            value = str(name.get(key) or "").strip()
            if value:
                return value

    # Для ИП/неполных ответов иногда удобнее брать value.
    value = str(suggestion.get("value") or "").strip()
    if value:
        return value

    return None


def lookup_legal_entity_name_by_dadata(buyer_inn: str, settings: dict[str, str]) -> Optional[str]:
    """
    Ищет название юрлица/ИП по ИНН через DaData Find Party API.

    Требуется DADATA_TOKEN в settings.txt.
    """
    token = str(settings.get("DADATA_TOKEN", "")).strip()
    if not token:
        print("Внимание: DADATA_TOKEN не задан, название юрлица по ИНН не загружено.")
        return None

    url = "https://suggestions.dadata.ru/suggestions/api/4_1/rs/findById/party"

    payload: dict[str, Any] = {
        "query": buyer_inn,
        "count": 1,
    }

    # Для 10-значного ИНН обычно ищем головную организацию.
    # Для 12-значного ИНН это ИП, branch_type не нужен.
    branch_type = str(settings.get("DADATA_BRANCH_TYPE", "MAIN")).strip().upper()
    if len(buyer_inn) == 10 and branch_type in {"MAIN", "BRANCH"}:
        payload["branch_type"] = branch_type

    if len(buyer_inn) == 12:
        payload["type"] = "INDIVIDUAL"
    elif len(buyer_inn) == 10:
        payload["type"] = "LEGAL"

    try:
        logging.info("REQUEST POST %s\n%s", url, json.dumps(payload, ensure_ascii=False, indent=2))
        response = requests.post(url, headers=dadata_headers(token), json=payload, timeout=15)
        logging.info("RESPONSE %s POST %s\n%s", response.status_code, url, response.text[:5000])
    except Exception as exc:
        print(f"Внимание: не удалось обратиться к DaData для поиска по ИНН: {exc}")
        logging.exception("DaData lookup failed")
        return None

    if response.status_code in {401, 403}:
        print("Внимание: DaData отклонила DADATA_TOKEN. Проверьте токен в settings.txt.")
        return None

    if response.status_code >= 400:
        print(f"Внимание: DaData вернула HTTP {response.status_code}. Название юрлица не загружено.")
        return None

    try:
        body = response.json()
    except Exception:
        print("Внимание: DaData вернула не JSON. Название юрлица не загружено.")
        return None

    suggestions = body.get("suggestions")
    if not isinstance(suggestions, list) or not suggestions:
        print(f"Внимание: DaData не нашла организацию/ИП по ИНН {buyer_inn}.")
        return None

    name = extract_dadata_party_name(suggestions[0])
    if name:
        print(f"Название покупателя найдено по ИНН: {name}")
    return name


def lookup_legal_entity_name_by_inn(buyer_inn: Optional[str], settings: dict[str, str]) -> Optional[str]:
    """
    Универсальная точка поиска названия юрлица/ИП по ИНН.

    Сейчас реализован провайдер DADATA.
    Контур.Фокус можно добавить отдельным провайдером при наличии API-ключа и выбранного тарифа.
    """
    if not buyer_inn:
        return None

    if not bool_setting(settings, "LEGAL_ENTITY_LOOKUP_ENABLED", True):
        return None

    provider = str(settings.get("LEGAL_ENTITY_LOOKUP_PROVIDER", "DADATA")).strip().upper()
    if provider in {"", "NONE", "OFF", "0", "DISABLED"}:
        return None

    if provider == "DADATA":
        return lookup_legal_entity_name_by_dadata(buyer_inn, settings)

    print(f"Внимание: неизвестный LEGAL_ENTITY_LOOKUP_PROVIDER = {provider}. Поиск по ИНН пропущен.")
    return None




def prompt_with_default(prompt: str, default: str = "") -> str:
    """Запрос значения с Enter-по-умолчанию."""
    if default:
        value = input(f"{prompt} [{default}]: ").strip()
        return value or default
    return input(f"{prompt}: ").strip()


def prompt_session_cashier(settings_path: Path, settings: dict[str, str]) -> str:
    current = str(settings.get("DEFAULT_CASHIER_NAME", "")).strip()
    while True:
        cashier_name = prompt_with_default("Введите имя кассира на текущую смену", current).strip()
        if cashier_name:
            settings["DEFAULT_CASHIER_NAME"] = cashier_name
            save_settings(settings_path, settings)
            print(f"Кассир текущего сеанса: {cashier_name}")
            return cashier_name
        print("Имя кассира не может быть пустым.")


def prompt_payment_type(settings_path: Path, settings: dict[str, str]) -> str:
    """
    Спрашивает тип оплаты.

    Для безнала конкретный способ приема выбирается настройкой:
    CASHLESS_PAYMENT_PROVIDER = EXTERNAL_TERMINAL / TBANK_SBP

    0 / q / отмена — отмена ввода и возврат в главное меню.
    """
    current = str(settings.get("DEFAULT_PAYMENT_TYPE", "CASHLESS")).strip().upper() or "CASHLESS"
    if current == "SBP":
        current = "CASHLESS"

    provider = normalize_cashless_payment_provider(settings.get("CASHLESS_PAYMENT_PROVIDER", "EXTERNAL_TERMINAL"))
    default_label = "Наличные" if current == "CASH" else "Безнал"

    print("\nТип оплаты:")
    print("  1. Наличные")
    print(f"  2. Безнал ({cashless_provider_label(provider)})")
    print("  0. Отмена и возврат в главное меню")
    print("\nСпособ безналичной оплаты меняется в настройках: CASHLESS_PAYMENT_PROVIDER.")
    while True:
        raw = input_or_cancel(f"Выберите тип оплаты [{default_label}]: ", stage="выбор типа оплаты")
        if not raw.strip():
            raw = default_label
        if raw.strip() == "1":
            raw = "Наличные"
        elif raw.strip() == "2":
            raw = "Безнал"

        try:
            payment_type = parse_payment_type(raw)
            if payment_type == "SBP":
                # Ручной ввод "СБП" оставляем как быстрый способ выбрать T-Банк.
                settings["DEFAULT_PAYMENT_TYPE"] = "CASHLESS"
                settings["CASHLESS_PAYMENT_PROVIDER"] = "TBANK_SBP"
                settings["TBANK_SBP_ENABLED"] = "1"
                save_settings(settings_path, settings)
                return "SBP"

            if payment_type == "CASHLESS":
                settings["DEFAULT_PAYMENT_TYPE"] = "CASHLESS"
                save_settings(settings_path, settings)
                provider = normalize_cashless_payment_provider(settings.get("CASHLESS_PAYMENT_PROVIDER", "EXTERNAL_TERMINAL"))
                return "SBP" if provider == "TBANK_SBP" else "CASHLESS"

            settings["DEFAULT_PAYMENT_TYPE"] = payment_type
            save_settings(settings_path, settings)
            return payment_type
        except Exception as exc:
            print(f"Ошибка: {exc}")

def prompt_email_contact() -> Optional[str]:
    while True:
        raw = input_or_cancel("Email покупателя [0 = отмена]: ", stage="ввод email покупателя").strip()
        if not raw:
            return None
        email, _phone = split_buyer_contact_optional(raw)
        if email:
            return email
        print("Email не распознан. Пример: client@example.com")


def prompt_phone_contact() -> Optional[str]:
    while True:
        raw = input_or_cancel("Телефон покупателя [0 = отмена]: ", stage="ввод телефона покупателя").strip()
        if not raw:
            return None
        _email, phone = split_buyer_contact_optional(raw)
        if phone:
            return phone
        print("Телефон не распознан. Пример: +79991234567 или 89991234567")


def prompt_buyer_contacts() -> tuple[Optional[str], Optional[str]]:
    """
    Сначала спрашивает, какой контакт покупателя вводить:
    email, телефон или оба варианта.

    Вариант "не указывать" удалён — для электронного чека оператор должен
    выбрать хотя бы один контакт покупателя.

    0 / q / отмена — отмена ввода и возврат в главное меню.
    """
    print("\nКонтакт покупателя для электронного чека:")
    print("  1. Email")
    print("  2. Телефон")
    print("  3. Email + телефон")
    print("  0. Отмена и возврат в главное меню")

    while True:
        choice = input_or_cancel("Выберите вариант [1/2/3/0]: ", stage="выбор контакта покупателя").strip().lower()

        if choice in {"1", "email", "e", "емайл", "почта"}:
            email = prompt_email_contact()
            if email:
                return email, None
            print("Email обязателен для выбранного варианта. Введите email или 0 для отмены.")

        elif choice in {"2", "телефон", "phone", "p", "т"}:
            phone = prompt_phone_contact()
            if phone:
                return None, phone
            print("Телефон обязателен для выбранного варианта. Введите телефон или 0 для отмены.")

        elif choice in {"3", "оба", "both", "email+phone", "email + phone", "e+p"}:
            email = prompt_email_contact()
            phone = prompt_phone_contact()
            if email and phone:
                return email, phone
            print("Для выбранного варианта нужно указать и email, и телефон. Введите данные или 0 для отмены.")

        else:
            print("Введите 1 для email, 2 для телефона, 3 для обоих вариантов или 0 для отмены.")

def prompt_buyer_type_and_legal(settings: dict[str, str]) -> tuple[str, Optional[str], Optional[str]]:
    print("\nТип покупателя:")
    print("  1. Физлицо")
    print("  2. Юрлицо / ИП")
    print("  0. Отмена и возврат в главное меню")
    while True:
        answer = input_or_cancel("Выберите тип покупателя [Enter = 1, 0 = отмена]: ", stage="выбор типа покупателя").strip().lower()
        if answer in {"", "1", "ф", "физ", "физлицо"}:
            return "INDIVIDUAL", None, None
        if answer in {"2", "ю", "юр", "юл", "юрлицо", "ип", "организация"}:
            break
        print("Введите 1 для физлица, 2 для юрлица / ИП или 0 для отмены.")

    while True:
        buyer_inn = normalize_buyer_inn(input_or_cancel("ИНН юрлица / ИП [0 = отмена]: ", stage="ввод ИНН юрлица / ИП").strip())
        if buyer_inn and len(buyer_inn) in {10, 12}:
            break
        print("ИНН должен содержать 10 цифр для юрлица или 12 цифр для ИП.")

    legal_name_raw = input_or_cancel("Наименование юрлица / ИП [Enter = найти по ИНН, 0 = отмена]: ", stage="ввод наименования юрлица / ИП").strip()
    legal_name = legal_name_raw or None
    if not legal_name:
        legal_name = lookup_legal_entity_name_by_inn(buyer_inn, settings)
    while not legal_name:
        legal_name = input_or_cancel("Введите наименование юрлица / ИП вручную [0 = отмена]: ", stage="ввод наименования юрлица / ИП").strip() or None
        if not legal_name:
            print("Для чека на юрлицо / ИП нужно название покупателя.")
    return "LEGAL_ENTITY", legal_name, buyer_inn

def prompt_tax_mode(settings_path: Path, settings: dict[str, str]) -> str:
    """
    Показывает список систем налогообложения на русском языке
    и принимает выбор цифрой.

    В Dreamkas API код OSN не используется — общая система отправляется как DEFAULT.
    Здесь 0 является рабочим значением DEFAULT, поэтому для отмены используйте q / отмена / выход.
    """
    current = str(settings.get("DEFAULT_TAX_MODE", "SIMPLE_WO")).strip().upper() or "SIMPLE_WO"
    current = normalize_tax_mode_for_dreamkas(current)
    current_label = tax_mode_label(current)

    print("\nСистема налогообложения:")
    print("  0. По умолчанию на кассе / общая система — DEFAULT")
    print("  1. Упрощённая система — доходы — УСН доходы")
    print("  2. Упрощённая система — доходы минус расходы — УСН доходы-расходы")
    print("  3. Единый сельскохозяйственный налог — ЕСХН")
    print("  4. Патентная система налогообложения — ПСН / патент")
    print("  q. Отмена и возврат в главное меню")
    print("\nВажно: Dreamkas API не принимает отдельный код OSN. Для общей системы используется DEFAULT.")
    print(f"\nТекущее значение по умолчанию: {current_label}")

    while True:
        raw = input_or_cancel(
            "Выберите систему налогообложения [0/1/2/3/4, Enter = текущее, q = отмена]: ",
            allow_zero_as_value=True,
            stage="выбор системы налогообложения",
        ).strip()
        if not raw:
            tax_mode = current
        else:
            try:
                tax_mode = parse_tax_mode(raw)
            except Exception as exc:
                print(f"Ошибка: {exc}")
                continue

        tax_mode = normalize_tax_mode_for_dreamkas(tax_mode)
        print(f"Выбрано: {tax_mode_label(tax_mode)}")
        settings["DEFAULT_TAX_MODE"] = tax_mode
        save_settings(settings_path, settings)
        return tax_mode

def prompt_sale_metadata(settings_path: Path, settings: dict[str, str], cashier_name: str) -> ReceiptInputMeta:
    print("\nЗаполнение реквизитов нового чека")
    print("-" * 72)
    payment_type = prompt_payment_type(settings_path, settings)
    buyer_email, buyer_phone = prompt_buyer_contacts()
    buyer_type, legal_name, buyer_inn = prompt_buyer_type_and_legal(settings)
    tax_mode = prompt_tax_mode(settings_path, settings)
    return ReceiptInputMeta(
        payment_type=payment_type,
        buyer_email=buyer_email,
        buyer_phone=buyer_phone,
        buyer_type=buyer_type,
        buyer_legal_name=legal_name,
        buyer_inn=buyer_inn,
        tax_mode=tax_mode,
        cashier_name=cashier_name,
    )


def read_buyer_legal_entity_template(ws: Any, settings: dict[str, str]) -> tuple[Optional[str], Optional[str], str, str, str, Optional[str], Optional[str], int]:
    """
    Шаблон v6.30:
      1 строка — тип оплаты
      2 строка — email покупателя
      3 строка — телефон покупателя
      4 строка — тип покупателя: Физлицо / Юрлицо
      5 строка — наименование юрлица / ИП
      6 строка — ИНН юрлица / ИП
      7 строка — СНО
      8 строка — кассир
      9 строка — шапка таблицы
      10 строка и ниже — позиции
    """
    buyer_email, buyer_phone = read_buyer_contacts_new_template(ws)
    legal_name = read_meta_cell(ws, 5).strip() or None
    buyer_inn = normalize_buyer_inn(read_meta_cell(ws, 6))
    buyer_type = normalize_buyer_type(read_meta_cell(ws, 4), legal_name, buyer_inn)

    # Если пользователь ввёл только ИНН, пробуем получить название юрлица/ИП автоматически.
    if buyer_type == "LEGAL_ENTITY" and buyer_inn and not legal_name:
        legal_name = lookup_legal_entity_name_by_inn(buyer_inn, settings)

        # Чтобы пользователь видел, что было подставлено, можно записать название обратно в Excel.
        # Это не критично: если файл открыт и заблокирован, просто пропускаем.
        if legal_name and bool_setting(settings, "DADATA_UPDATE_EXCEL_NAME", True):
            try:
                ws.cell(row=5, column=2).value = legal_name
            except Exception:
                pass

    tax_mode = parse_tax_mode(read_meta_cell(ws, 7))
    cashier_name = read_meta_cell(ws, 8) or "Не указан"
    return buyer_email, buyer_phone, buyer_type, tax_mode, cashier_name, legal_name, buyer_inn, 10


def build_legal_entity_ffd_tags(
    buyer_type: str,
    legal_name: Optional[str],
    buyer_inn: Optional[str],
    settings: dict[str, str],
) -> list[dict[str, Any]]:
    """
    Для расчётов с юрлицом / ИП добавляем ФФД-теги:
      1227 — покупатель (клиент)
      1228 — ИНН покупателя (клиента)

    Dreamkas /api/receipts поддерживает массив tags с номером тега и значением.
    Если конкретная касса/прошивка не примет строковые теги, можно отключить отправку:
      SEND_LEGAL_ENTITY_TAGS_TO_DREAMKAS = 0
    """
    if buyer_type != "LEGAL_ENTITY":
        return []
    if str(settings.get("SEND_LEGAL_ENTITY_TAGS_TO_DREAMKAS", "1")).strip() in {"0", "false", "False", "нет", "Нет"}:
        return []

    tags: list[dict[str, Any]] = []
    if legal_name:
        tags.append({"tag": 1227, "value": str(legal_name)})
    if buyer_inn:
        tags.append({"tag": 1228, "value": str(buyer_inn)})
    return tags

def build_cashier_payload(cashier_name: str) -> dict[str, str]:
    """
    Dreamkas /api/receipts currently rejects the top-level "cashier" field.

    The cashier name from Excel is still stored locally:
    - precheck
    - SQLite
    - TXT/PDF receipt files

    But it is NOT sent to Dreamkas JSON, because the API returns:
    E_VALIDATION_OBJECT_ALLOW_UNKNOWN: "cashier" is not allowed.

    If Dreamkas later adds a supported cashier parameter for /receipts,
    this function can be updated.
    """
    return {}


def read_excel_receipt(excel_path: Path, settings: dict[str, str], meta: Optional[ReceiptInputMeta] = None) -> ReceiptDraft:
    excel_path = ensure_excel_template_available(excel_path)

    wb = load_workbook(filename=excel_path, data_only=True)
    ws = wb.active

    if meta is not None:
        payment_type = meta.payment_type
        buyer_email = meta.buyer_email
        buyer_phone = meta.buyer_phone
        buyer_type = meta.buyer_type
        buyer_legal_name = meta.buyer_legal_name
        buyer_inn = meta.buyer_inn
        tax_mode = meta.tax_mode
        cashier_name = meta.cashier_name
        first_position_row = 2 if looks_like_header_row(ws, 1) else 1
    else:
        # Совместимость со старыми Excel-шаблонами.
        payment_type = parse_payment_type(read_meta_cell(ws, 1))
        buyer_type = "INDIVIDUAL"
        buyer_legal_name: Optional[str] = None
        buyer_inn: Optional[str] = None
        if looks_like_header_row(ws, 9):
            (
                buyer_email, buyer_phone, buyer_type, tax_mode, cashier_name,
                buyer_legal_name, buyer_inn, first_position_row,
            ) = read_buyer_legal_entity_template(ws, settings)
            if buyer_legal_name and bool_setting(settings, "DADATA_UPDATE_EXCEL_NAME", True):
                try:
                    wb.save(excel_path)
                except Exception as exc:
                    print(f"Внимание: не удалось записать найденное название в Excel: {exc}")
        elif looks_like_header_row(ws, 6):
            buyer_email, buyer_phone = read_buyer_contacts_new_template(ws)
            tax_mode = parse_tax_mode(read_meta_cell(ws, 4))
            cashier_name = read_meta_cell(ws, 5) or "Не указан"
            first_position_row = 7
        elif looks_like_header_row(ws, 1):
            raise ValueError("Excel содержит только позиции. Реквизиты чека должны быть введены через терминал.")
        else:
            buyer_email, buyer_phone = split_buyer_contact(read_meta_cell(ws, 2))
            tax_mode = parse_tax_mode(read_meta_cell(ws, 3))
            cashier_name = read_meta_cell(ws, 4) or "Не указан"
            first_position_row = 6

    positions: list[ReceiptPosition] = []
    for row in range(first_position_row, ws.max_row + 1):
        name_raw = ws.cell(row=row, column=1).value
        if name_raw is None or str(name_raw).strip() == "":
            # Считаем первую пустую строку окончанием списка позиций.
            break

        name = str(name_raw).strip()
        item_type = parse_item_type(ws.cell(row=row, column=2).value, row)
        quantity = normalize_quantity(ws.cell(row=row, column=3).value, row)
        price_kopecks = rub_to_kopecks(ws.cell(row=row, column=4).value, row)
        price_sum_kopecks = int((Decimal(price_kopecks) * quantity).quantize(Decimal("1"), rounding=ROUND_HALF_UP))
        tax = parse_tax(ws.cell(row=row, column=5).value, row)

        positions.append(
            ReceiptPosition(
                name=name,
                item_type=item_type,
                quantity=quantity,
                price_kopecks=price_kopecks,
                price_sum_kopecks=price_sum_kopecks,
                tax=tax,
            )
        )

    if not positions:
        raise ValueError(f"В Excel нет товарных позиций. Позиции должны начинаться с {first_position_row}-й строки.")
    if len(positions) > 100:
        raise ValueError("В одном чеке Dreamkas допускает максимум 100 позиций")

    total_kopecks = sum(p.price_sum_kopecks for p in positions)
    if total_kopecks <= 0:
        raise ValueError("Сумма чека должна быть больше 0")

    external_id = str(uuid.uuid4())
    device_id = int(get_required(settings, "DEVICE_ID"))
    shop_id = int(get_required(settings, "SHOP_ID"))
    timeout_minutes = max(5, get_int(settings, "TIMEOUT_MINUTES", 5))
    operation_type = settings.get("OPERATION_TYPE", "SALE").strip().upper() or "SALE"

    if operation_type not in {"SALE", "REFUND", "OUTFLOW", "OUTFLOW_REFUND"}:
        raise ValueError("OPERATION_TYPE должен быть SALE, REFUND, OUTFLOW или OUTFLOW_REFUND")

    attributes: dict[str, str] = {}
    if buyer_email:
        attributes["email"] = buyer_email
    if buyer_phone:
        attributes["phone"] = buyer_phone

    payload_positions: list[dict[str, Any]] = []
    for p in positions:
        payload_positions.append(
            {
                "name": p.name,
                "type": p.item_type,
                "quantity": decimal_to_json_number(p.quantity),
                "price": p.price_kopecks,
                "priceSum": p.price_sum_kopecks,
                "tax": p.tax,
            }
        )

    payload = {
        "externalId": external_id,
        "deviceId": device_id,
        "shopId": shop_id,
        "type": operation_type,
        "timeout": timeout_minutes,
        "taxMode": normalize_tax_mode_for_dreamkas(tax_mode),
        "positions": payload_positions,
        "payments": [
            {
                "sum": total_kopecks,
                # Для Dreamkas платеж по СБП является безналичным.
                "type": "CASHLESS" if payment_type == "SBP" else payment_type,
            }
        ],
        "attributes": attributes,
        "total": {
            "priceSum": total_kopecks,
        },
    }

    legal_entity_tags = build_legal_entity_ffd_tags(buyer_type, buyer_legal_name, buyer_inn, settings)
    if legal_entity_tags:
        existing_tags = payload.get("tags")
        if isinstance(existing_tags, list):
            payload["tags"] = existing_tags + legal_entity_tags
        else:
            payload["tags"] = legal_entity_tags

    cashier_payload = build_cashier_payload(cashier_name)
    if cashier_payload:
        payload["cashier"] = cashier_payload

    return ReceiptDraft(
        external_id=external_id,
        payment_type=payment_type,
        buyer_email=buyer_email,
        buyer_phone=buyer_phone,
        buyer_type=buyer_type,
        buyer_legal_name=buyer_legal_name,
        buyer_inn=buyer_inn,
        tax_mode=tax_mode,
        cashier_name=cashier_name,
        positions=positions,
        total_kopecks=total_kopecks,
        payload=payload,
    )


def clear_excel_template_after_success(excel_path: Path) -> None:
    """
    Очищает только товарные позиции.

    Реквизиты чека вводятся в терминале, поэтому в новом Excel-шаблоне
    нет строк оплаты, покупателя, СНО и кассира.
    """
    excel_path = ensure_excel_template_available(excel_path)
    wb = load_workbook(filename=excel_path)
    ws = wb.active
    first_position_row = 2 if looks_like_header_row(ws, 1) else 1
    max_row_to_clear = max(ws.max_row, first_position_row + 150)
    for row in range(first_position_row, max_row_to_clear + 1):
        for col in range(1, 6):
            ws.cell(row=row, column=col).value = None
    wb.save(excel_path)

def format_money(kopecks: int) -> str:
    return f"{kopecks / 100:.2f} ₽"


def print_precheck(draft: ReceiptDraft) -> None:
    print("\n" + "=" * 72)
    print("ПРЕДЧЕК ДЛЯ ПРОВЕРКИ")
    print("=" * 72)
    print(f"Тип операции:        {draft.payload.get('type')}")
    print(f"Тип оплаты:          {payment_label(draft.payment_type)}")
    print(f"Покупатель email:    {draft.buyer_email or '-'}")
    print(f"Покупатель телефон:  {draft.buyer_phone or '-'}")
    print(f"Тип покупателя:      {'Юрлицо / ИП' if draft.buyer_type == 'LEGAL_ENTITY' else 'Физлицо'}")
    if draft.buyer_type == 'LEGAL_ENTITY':
        print(f"Покупатель юрлицо:   {draft.buyer_legal_name or '-'}")
        print(f"ИНН покупателя:      {draft.buyer_inn or '-'}")
    print(f"СНО:                 {tax_mode_label(draft.tax_mode)}")
    print(f"Кассир:              {draft.cashier_name}")
    print(f"External ID:         {draft.external_id}")
    if draft.parent_external_id:
        print(f"Исходный чек:        {draft.parent_external_id}")
    if draft.parent_receipt_id:
        print(f"Исходный Receipt ID: {draft.parent_receipt_id}")
    if draft.refund_mode:
        print(f"Режим возврата:      {draft.refund_mode}")
    print("-" * 72)
    print(f"{'№':>2}  {'Наименование':<32} {'Тип':<9} {'Кол-во':>8} {'Цена':>12} {'Сумма':>12} {'НДС':>12}")
    print("-" * 72)
    for i, p in enumerate(draft.positions, start=1):
        name = p.name[:32]
        qty = str(p.quantity.normalize())
        print(
            f"{i:>2}  {name:<32} {p.item_type:<9} {qty:>8} "
            f"{format_money(p.price_kopecks):>12} {format_money(p.price_sum_kopecks):>12} {p.tax:>12}"
        )
    print("-" * 72)
    print(f"ИТОГО: {format_money(draft.total_kopecks)}")
    print("=" * 72)
    print("Внимание: после отправки касса может фискализировать чек. Проверь сумму и позиции.")


def init_db(db_path: Path) -> sqlite3.Connection:
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS receipts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            external_id TEXT UNIQUE NOT NULL,
            operation_id TEXT,
            receipt_id TEXT,
            status TEXT NOT NULL,
            created_at TEXT NOT NULL,
            sent_at TEXT,
            completed_at TEXT,
            payment_type TEXT,
            buyer_email TEXT,
            buyer_phone TEXT,
            buyer_type TEXT,
            buyer_legal_name TEXT,
            buyer_inn TEXT,
            tax_mode TEXT,
            cashier_name TEXT,
            amount_kopecks INTEGER,
            receipt_type TEXT,
            parent_external_id TEXT,
            parent_receipt_id TEXT,
            refund_mode TEXT,
            refunded_positions_json TEXT,
            request_json TEXT,
            operation_json TEXT,
            receipt_json TEXT,
            error_code TEXT,
            error_message TEXT,
            receipt_txt_path TEXT,
            qr_path TEXT,
            pdf_path TEXT,
            payload_hash TEXT,
            precheck_hash TEXT,
            precheck_json TEXT,
            excel_file_path TEXT,
            sbp_payment_id TEXT,
            sbp_order_id TEXT,
            sbp_status TEXT,
            sbp_qr_payload TEXT,
            sbp_qr_path TEXT,
            sbp_json TEXT
        )
        """
    )
    ensure_db_columns(conn)
    conn.commit()
    return conn


def ensure_db_columns(conn: sqlite3.Connection) -> None:
    """Мягко добавляет новые колонки в уже существующую SQLite-базу."""
    existing = {row[1] for row in conn.execute("PRAGMA table_info(receipts)").fetchall()}
    required = {
        "receipt_type": "TEXT",
        "parent_external_id": "TEXT",
        "parent_receipt_id": "TEXT",
        "refund_mode": "TEXT",
        "refunded_positions_json": "TEXT",
        "pdf_path": "TEXT",
        "payload_hash": "TEXT",
        "precheck_hash": "TEXT",
        "precheck_json": "TEXT",
        "excel_file_path": "TEXT",
        "buyer_type": "TEXT",
        "buyer_legal_name": "TEXT",
        "buyer_inn": "TEXT",
        "sbp_payment_id": "TEXT",
        "sbp_order_id": "TEXT",
        "sbp_status": "TEXT",
        "sbp_qr_payload": "TEXT",
        "sbp_qr_path": "TEXT",
        "sbp_json": "TEXT",
    }
    for column, column_type in required.items():
        if column not in existing:
            conn.execute(f"ALTER TABLE receipts ADD COLUMN {column} {column_type}")


def payload_type_from_json(request_json: Optional[str]) -> str:
    if not request_json:
        return ""
    try:
        payload = json.loads(request_json)
    except Exception:
        return ""
    if isinstance(payload, dict):
        return str(payload.get("type") or "").upper()
    return ""



def payload_for_hash(payload: dict[str, Any]) -> dict[str, Any]:
    data = deepcopy(payload)
    data.pop("externalId", None)
    # timeout не влияет на содержимое чека, только на ожидание выполнения.
    data.pop("timeout", None)
    return data


def receipt_payload_hash(payload: dict[str, Any]) -> str:
    raw = json.dumps(payload_for_hash(payload), ensure_ascii=False, sort_keys=True, default=str)
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def position_for_precheck(position: ReceiptPosition) -> dict[str, Any]:
    return {
        "name": position.name,
        "type": position.item_type,
        "quantity": decimal_to_json_number(position.quantity),
        "price_kopecks": position.price_kopecks,
        "price_rub": float(Decimal(position.price_kopecks) / Decimal("100")),
        "sum_kopecks": position.price_sum_kopecks,
        "sum_rub": float(Decimal(position.price_sum_kopecks) / Decimal("100")),
        "tax": position.tax,
    }


def precheck_to_dict(draft: ReceiptDraft, excel_path: Optional[Path] = None) -> dict[str, Any]:
    """
    Полный слепок предчека до отправки в Dreamkas.

    Его сохраняем в SQLite, чтобы потом можно было проверить, что именно
    пользователь видел и подтверждал перед фискализацией.
    """
    data: dict[str, Any] = {
        "app_version": APP_VERSION,
        "created_at": datetime.now().isoformat(timespec="seconds"),
        "external_id": draft.external_id,
        "operation_type": str(draft.payload.get("type") or ""),
        "payment_type": draft.payment_type,
        "buyer_email": draft.buyer_email,
        "buyer_phone": draft.buyer_phone,
        "buyer_type": draft.buyer_type,
        "buyer_legal_name": draft.buyer_legal_name,
        "buyer_inn": draft.buyer_inn,
        "tax_mode": draft.tax_mode,
        "cashier_name": draft.cashier_name,
        "device_id": draft.payload.get("deviceId"),
        "shop_id": draft.payload.get("shopId"),
        "total_kopecks": draft.total_kopecks,
        "total_rub": float(Decimal(draft.total_kopecks) / Decimal("100")),
        "positions": [position_for_precheck(p) for p in draft.positions],
        "parent_external_id": draft.parent_external_id,
        "parent_receipt_id": draft.parent_receipt_id,
        "refund_mode": draft.refund_mode,
        "refunded_positions": draft.refunded_positions or [],
    }
    if excel_path is not None:
        data["excel_file_path"] = str(excel_path)
    return data


def precheck_for_hash(draft: ReceiptDraft) -> dict[str, Any]:
    """
    Данные для поиска дублей.

    Не включаем external_id и дату, потому что они всегда разные.
    Включаем содержимое чека: тип операции, покупателя, оплату, СНО, кассира,
    позиции и сумму. Благодаря этому дубль ловится даже если Excel-шаблон
    перезаписали или открыли заново.
    """
    return {
        "operation_type": str(draft.payload.get("type") or ""),
        "payment_type": draft.payment_type,
        "buyer_email": draft.buyer_email,
        "buyer_phone": draft.buyer_phone,
        "buyer_type": draft.buyer_type,
        "buyer_legal_name": draft.buyer_legal_name,
        "buyer_inn": draft.buyer_inn,
        "tax_mode": draft.tax_mode,
        "cashier_name": draft.cashier_name,
        "total_kopecks": draft.total_kopecks,
        "positions": [position_for_precheck(p) for p in draft.positions],
        "parent_external_id": draft.parent_external_id,
        "parent_receipt_id": draft.parent_receipt_id,
        "refund_mode": draft.refund_mode,
        "refunded_positions": draft.refunded_positions or [],
    }


def receipt_precheck_hash(draft: ReceiptDraft) -> str:
    raw = json.dumps(precheck_for_hash(draft), ensure_ascii=False, sort_keys=True, default=str)
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def validate_draft_for_send(draft: ReceiptDraft) -> tuple[list[str], list[str]]:
    errors: list[str] = []
    warnings: list[str] = []

    if not draft.buyer_email and not draft.buyer_phone:
        warnings.append("Email/телефон покупателя не указан. Электронная отправка чека покупателю будет недоступна.")
    if draft.buyer_email and "@" not in draft.buyer_email:
        errors.append(f"Email покупателя выглядит неверно: {draft.buyer_email}")
    if draft.buyer_phone:
        digits = re.sub(r"\D", "", draft.buyer_phone)
        if len(digits) < 10:
            errors.append(f"Телефон покупателя выглядит неверно: {draft.buyer_phone}")
    if draft.buyer_type == "LEGAL_ENTITY":
        if not draft.buyer_inn:
            errors.append("Для чека на юрлицо / ИП нужно заполнить ИНН покупателя.")
        elif len(re.sub(r"\D", "", draft.buyer_inn)) not in {10, 12}:
            errors.append(f"ИНН покупателя должен содержать 10 или 12 цифр: {draft.buyer_inn}")

        if not draft.buyer_legal_name:
            errors.append(
                "Не удалось определить наименование юрлица / ИП по ИНН. "
                "Введите наименование юрлица / ИП вручную или настройте DADATA_TOKEN."
            )

        if not any(tag.get("tag") == 1228 for tag in draft.payload.get("tags", []) if isinstance(tag, dict)):
            warnings.append("Данные юрлица сохраняются локально, но ФФД-теги 1227/1228 не отправляются в Dreamkas. Проверь настройку SEND_LEGAL_ENTITY_TAGS_TO_DREAMKAS.")
    if not draft.cashier_name or draft.cashier_name.strip().lower() in {"не указан", "admin", "administrator", "администратор"}:
        warnings.append("Имя кассира пустое или похоже на значение по умолчанию. В Excel лучше указать реальное имя кассира.")
    if not draft.positions:
        errors.append("В чеке нет позиций.")
    if len(draft.positions) > 100:
        errors.append("В одном чеке не должно быть больше 100 позиций.")
    if draft.total_kopecks <= 0:
        errors.append("Сумма чека должна быть больше 0.")

    seen_names: set[str] = set()
    for index, p in enumerate(draft.positions, start=1):
        if not p.name.strip():
            errors.append(f"Позиция {index}: пустое наименование.")
        if p.name in seen_names:
            warnings.append(f"Позиция {index}: наименование повторяется: {p.name}")
        seen_names.add(p.name)
        if p.quantity <= 0:
            errors.append(f"Позиция {index}: количество должно быть больше 0.")
        if p.price_kopecks <= 0:
            errors.append(f"Позиция {index}: цена должна быть больше 0.")
        if p.price_sum_kopecks <= 0:
            errors.append(f"Позиция {index}: сумма должна быть больше 0.")
        if p.tax not in {
            "NDS_NO_TAX", "NDS_0", "NDS_5", "NDS_7", "NDS_10", "NDS_20", "NDS_22",
            "NDS_5_CALCULATED", "NDS_7_CALCULATED", "NDS_10_CALCULATED", "NDS_20_CALCULATED", "NDS_22_CALCULATED",
            "NDS_10_110", "NDS_20_120",
        }:
            errors.append(f"Позиция {index}: неизвестная ставка НДС: {p.tax}")
        if p.item_type not in {"COUNTABLE", "SERVICE"}:
            errors.append(f"Позиция {index}: неизвестный тип позиции: {p.item_type}")

    return errors, warnings


def print_validation_result(errors: list[str], warnings: list[str]) -> bool:
    if errors:
        print("\nОШИБКИ В ЧЕКЕ - отправка заблокирована:")
        for item in errors:
            print(f"  - {item}")
    if warnings:
        print("\nПРЕДУПРЕЖДЕНИЯ:")
        for item in warnings:
            print(f"  - {item}")
    if errors:
        return False
    if warnings:
        answer = input("\nПродолжить несмотря на предупреждения? Введите ДА: ").strip().lower()
        return answer in {"да", "д", "yes", "y"}
    return True


def recent_duplicate_rows(conn: sqlite3.Connection, draft: ReceiptDraft, settings: dict[str, str]) -> list[sqlite3.Row]:
    precheck_hash = receipt_precheck_hash(draft)

    # Новый режим: ищем дубли по предчеку за N дней.
    # Старый DUPLICATE_WINDOW_MINUTES оставлен для совместимости, но если
    # DUPLICATE_LOOKBACK_DAYS задан, он приоритетнее.
    days_raw = str(settings.get("DUPLICATE_LOOKBACK_DAYS", "30")).strip()
    if days_raw:
        try:
            days = max(1, int(days_raw))
            cutoff = (datetime.now() - timedelta(days=days)).isoformat(timespec="seconds")
            period_text = f"{days} дн."
        except ValueError:
            cutoff = (datetime.now() - timedelta(days=30)).isoformat(timespec="seconds")
            period_text = "30 дн."
    else:
        minutes = max(1, get_int(settings, "DUPLICATE_WINDOW_MINUTES", 10))
        cutoff = (datetime.now() - timedelta(minutes=minutes)).isoformat(timespec="seconds")
        period_text = f"{minutes} мин."

    rows = conn.execute(
        """
        SELECT * FROM receipts
        WHERE precheck_hash = ?
          AND created_at >= ?
          AND status IN ('PENDING', 'IN_PROGRESS', 'SUCCESS')
        ORDER BY created_at DESC
        LIMIT 10
        """,
        (precheck_hash, cutoff),
    ).fetchall()

    # Совместимость со старыми записями, где precheck_hash еще не было.
    if rows:
        return rows

    payload_hash = receipt_payload_hash(draft.payload)
    return conn.execute(
        """
        SELECT * FROM receipts
        WHERE payload_hash = ?
          AND created_at >= ?
          AND status IN ('PENDING', 'IN_PROGRESS', 'SUCCESS')
        ORDER BY created_at DESC
        LIMIT 10
        """,
        (payload_hash, cutoff),
    ).fetchall()


def print_duplicate_details(row: sqlite3.Row) -> None:
    print(
        f"  DB ID {row['id']} | {row['created_at']} | "
        f"{row['status']} | {format_money(int(row['amount_kopecks'] or 0))} | "
        f"receiptId: {row['receipt_id'] or '-'}"
    )
    precheck_raw = row['precheck_json'] if 'precheck_json' in row.keys() else None
    if precheck_raw:
        try:
            precheck = json.loads(precheck_raw)
            print(f"     Покупатель: {precheck.get('buyer_email') or '-'} / {precheck.get('buyer_phone') or '-'}")
            print(f"     Кассир: {precheck.get('cashier_name') or '-'}")
            print(f"     Позиций: {len(precheck.get('positions') or [])}")
            for index, pos in enumerate((precheck.get('positions') or [])[:5], start=1):
                print(
                    f"       {index}. {pos.get('name')} | "
                    f"{pos.get('quantity')} x {pos.get('price_rub')} = {pos.get('sum_rub')}"
                )
            if len(precheck.get('positions') or []) > 5:
                print("       ...")
        except Exception:
            pass


def confirm_no_recent_duplicate(conn: sqlite3.Connection, draft: ReceiptDraft, settings: dict[str, str]) -> bool:
    rows = recent_duplicate_rows(conn, draft, settings)
    if not rows:
        return True
    print("\nВНИМАНИЕ: найден похожий предчек в SQLite.")
    print("Есть подозрение, что такой чек уже выбивался или находится в обработке.")
    print("Найденные записи:")
    for row in rows:
        print_duplicate_details(row)
    answer = input("Все равно пробивать этот чек? Введите ДА: ").strip().lower()
    return answer in {"да", "д", "yes", "y"}


def db_insert_draft(conn: sqlite3.Connection, draft: ReceiptDraft, excel_path: Optional[Path] = None) -> None:
    receipt_type = str(draft.payload.get("type") or "").upper()
    precheck_json = precheck_to_dict(draft, excel_path)
    conn.execute(
        """
        INSERT INTO receipts (
            external_id, status, created_at, payment_type, buyer_email, buyer_phone,
            buyer_type, buyer_legal_name, buyer_inn,
            tax_mode, cashier_name, amount_kopecks, receipt_type, parent_external_id,
            parent_receipt_id, refund_mode, refunded_positions_json, request_json, payload_hash,
            precheck_hash, precheck_json, excel_file_path
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            draft.external_id,
            "draft",
            datetime.now().isoformat(timespec="seconds"),
            draft.payment_type,
            draft.buyer_email,
            draft.buyer_phone,
            draft.buyer_type,
            draft.buyer_legal_name,
            draft.buyer_inn,
            draft.tax_mode,
            draft.cashier_name,
            draft.total_kopecks,
            receipt_type,
            draft.parent_external_id,
            draft.parent_receipt_id,
            draft.refund_mode,
            json.dumps(draft.refunded_positions or [], ensure_ascii=False),
            json.dumps(draft.payload, ensure_ascii=False, indent=2),
            receipt_payload_hash(draft.payload),
            receipt_precheck_hash(draft),
            json.dumps(precheck_json, ensure_ascii=False, indent=2),
            str(excel_path) if excel_path is not None else None,
        ),
    )
    conn.commit()


def db_update(conn: sqlite3.Connection, external_id: str, **fields: Any) -> None:
    if not fields:
        return
    columns = ", ".join([f"{key} = ?" for key in fields])
    values = list(fields.values()) + [external_id]
    conn.execute(f"UPDATE receipts SET {columns} WHERE external_id = ?", values)
    conn.commit()


def dreamkas_headers(token: str) -> dict[str, str]:
    return {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json",
        "Accept": "application/json",
        "User-Agent": "dreamkas-receipt-python/1.2",
    }


def parse_api_response(response: requests.Response) -> Any:
    text = response.text.strip()
    if not text:
        return None
    try:
        return response.json()
    except ValueError:
        return {"raw": text}


def api_request(
    session: requests.Session,
    method: str,
    url: str,
    *,
    timeout_seconds: int,
    json_body: Optional[dict[str, Any]] = None,
) -> Any:
    log_json(f"REQUEST {method} {url}", json_body or {})
    try:
        response = session.request(method, url, json=json_body, timeout=timeout_seconds)
    except requests.RequestException as exc:
        logging.exception("Network error for %s %s", method, url)
        raise DreamkasError(f"Сетевая ошибка при запросе {method} {url}: {exc}") from exc

    body = parse_api_response(response)
    log_json(f"RESPONSE {response.status_code} {method} {url}", body)
    if response.status_code < 200 or response.status_code >= 300:
        message = None
        if isinstance(body, dict):
            message = body.get("message") or body.get("error") or body.get("code")
        raise DreamkasError(f"Dreamkas API вернул HTTP {response.status_code}: {message or body}")
    return body




def get_shops(
    session: requests.Session,
    api_base_url: str,
    timeout_seconds: int,
) -> list[dict[str, Any]]:
    """Получает список магазинов Dreamkas через GET /shops."""
    url = api_base_url.rstrip("/") + "/shops"
    body = api_request(session, "GET", url, timeout_seconds=timeout_seconds)
    if not isinstance(body, list):
        raise DreamkasError("Некорректный ответ Dreamkas при получении списка магазинов")
    return [item for item in body if isinstance(item, dict)]


def get_devices(
    session: requests.Session,
    api_base_url: str,
    timeout_seconds: int,
) -> list[dict[str, Any]]:
    """Получает список касс/устройств Dreamkas через GET /devices."""
    url = api_base_url.rstrip("/") + "/devices"
    body = api_request(session, "GET", url, timeout_seconds=timeout_seconds)
    if not isinstance(body, list):
        raise DreamkasError("Некорректный ответ Dreamkas при получении списка касс")
    return [item for item in body if isinstance(item, dict)]


def safe_int(value: Any) -> Optional[int]:
    """Пробует привести значение к int. Возвращает None, если привести нельзя."""
    if value is None:
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def choose_shop(shops: list[dict[str, Any]], current_shop_id: Optional[int]) -> dict[str, Any]:
    """Показывает список магазинов и возвращает выбранный магазин."""
    if not shops:
        raise DreamkasError("В Dreamkas API не найдено ни одного магазина")

    shops_sorted = sorted(shops, key=lambda item: (str(item.get("name") or ""), safe_int(item.get("id")) or 0))

    print("\nДоступные магазины Dreamkas:")
    print("-" * 72)
    for index, shop in enumerate(shops_sorted, start=1):
        shop_id = shop.get("id")
        name = shop.get("name") or "Без названия"
        current_mark = "  [текущий]" if current_shop_id is not None and safe_int(shop_id) == current_shop_id else ""
        print(f"{index:>2}. SHOP_ID = {shop_id} | {name}{current_mark}")
    print("-" * 72)

    default_index: Optional[int] = None
    if current_shop_id is not None:
        for index, shop in enumerate(shops_sorted, start=1):
            if safe_int(shop.get("id")) == current_shop_id:
                default_index = index
                break
    if default_index is None and len(shops_sorted) == 1:
        default_index = 1

    while True:
        if default_index:
            answer = input(f"Выберите магазин [Enter = {default_index}]: ").strip()
            if not answer:
                return shops_sorted[default_index - 1]
        else:
            answer = input("Выберите магазин по номеру: ").strip()

        if answer.isdigit():
            number = int(answer)
            if 1 <= number <= len(shops_sorted):
                return shops_sorted[number - 1]
        print("Некорректный выбор. Введите номер из списка.")


def choose_device(devices: list[dict[str, Any]], selected_shop_id: int, current_device_id: Optional[int]) -> dict[str, Any]:
    """Показывает кассы выбранного магазина и возвращает выбранную кассу."""
    filtered = [device for device in devices if safe_int(device.get("groupId")) == selected_shop_id]
    if not filtered:
        print("\nВнимание: у выбранного магазина не найдено касс по groupId.")
        print("Покажу все кассы из кабинета, но лучше проверить привязку кассы к магазину в Dreamkas.")
        filtered = devices

    if not filtered:
        raise DreamkasError("В Dreamkas API не найдено ни одной кассы/устройства")

    devices_sorted = sorted(filtered, key=lambda item: (str(item.get("name") or ""), safe_int(item.get("id")) or 0))

    print("\nДоступные кассы Dreamkas:")
    print("-" * 72)
    for index, device in enumerate(devices_sorted, start=1):
        device_id = device.get("id")
        name = device.get("name") or "Без названия"
        group_id = device.get("groupId")
        current_mark = "  [текущая]" if current_device_id is not None and safe_int(device_id) == current_device_id else ""
        print(f"{index:>2}. DEVICE_ID = {device_id} | {name} | SHOP_ID/groupId = {group_id}{current_mark}")
    print("-" * 72)

    default_index: Optional[int] = None
    if current_device_id is not None:
        for index, device in enumerate(devices_sorted, start=1):
            if safe_int(device.get("id")) == current_device_id:
                default_index = index
                break
    if default_index is None and len(devices_sorted) == 1:
        default_index = 1

    while True:
        if default_index:
            answer = input(f"Выберите кассу [Enter = {default_index}]: ").strip()
            if not answer:
                return devices_sorted[default_index - 1]
        else:
            answer = input("Выберите кассу по номеру: ").strip()

        if answer.isdigit():
            number = int(answer)
            if 1 <= number <= len(devices_sorted):
                return devices_sorted[number - 1]
        print("Некорректный выбор. Введите номер из списка.")


def select_shop_and_device(
    settings_path: Path,
    settings: dict[str, str],
    session: requests.Session,
    api_base_url: str,
    timeout_seconds: int,
) -> dict[str, str]:
    """
    Каждый запуск получает актуальные магазины/кассы из Dreamkas,
    дает пользователю выбрать магазин и кассу, затем сохраняет SHOP_ID / DEVICE_ID.
    """
    current_shop_id = safe_int(settings.get("SHOP_ID"))
    current_device_id = safe_int(settings.get("DEVICE_ID"))

    print("\nЗагружаю список магазинов из Dreamkas...")
    shops = get_shops(session, api_base_url, timeout_seconds)
    selected_shop = choose_shop(shops, current_shop_id)
    selected_shop_id = safe_int(selected_shop.get("id"))
    if selected_shop_id is None:
        raise DreamkasError("У выбранного магазина нет корректного id")

    print("\nЗагружаю список касс из Dreamkas...")
    devices = get_devices(session, api_base_url, timeout_seconds)
    selected_device = choose_device(devices, selected_shop_id, current_device_id)
    selected_device_id = safe_int(selected_device.get("id"))
    if selected_device_id is None:
        raise DreamkasError("У выбранной кассы нет корректного id")

    updated = dict(settings)
    updated["SHOP_ID"] = str(selected_shop_id)
    updated["SHOP_NAME"] = str(selected_shop.get("name") or "")
    updated["DEVICE_ID"] = str(selected_device_id)
    updated["DEVICE_NAME"] = str(selected_device.get("name") or "")
    save_settings(settings_path, updated)

    print("\nВыбор сохранен в settings.txt:")
    print(f"  SHOP_ID   = {updated['SHOP_ID']} | {updated.get('SHOP_NAME') or 'Без названия'}")
    print(f"  DEVICE_ID = {updated['DEVICE_ID']} | {updated.get('DEVICE_NAME') or 'Без названия'}")

    return updated

def send_receipt(
    session: requests.Session,
    api_base_url: str,
    draft: ReceiptDraft,
    timeout_seconds: int,
) -> dict[str, Any]:
    url = api_base_url.rstrip("/") + "/receipts"

    # Defensive cleanup: Dreamkas /api/receipts rejects top-level "cashier".
    # We keep cashier_name locally in SQLite/TXT/PDF, but do not send it to the API.
    payload = dict(draft.payload)
    payload.pop("cashier", None)

    body = api_request(session, "POST", url, timeout_seconds=timeout_seconds, json_body=payload)
    if not isinstance(body, dict):
        raise DreamkasError("Некорректный ответ Dreamkas при создании чека")
    return body


def get_operation(
    session: requests.Session,
    api_base_url: str,
    operation_id: str,
    timeout_seconds: int,
) -> dict[str, Any]:
    url = api_base_url.rstrip("/") + f"/operations/{operation_id}"
    body = api_request(session, "GET", url, timeout_seconds=timeout_seconds)
    if not isinstance(body, dict):
        raise DreamkasError("Некорректный ответ Dreamkas при проверке операции")
    return body


def get_receipt(
    session: requests.Session,
    api_base_url: str,
    receipt_id: str,
    timeout_seconds: int,
) -> dict[str, Any]:
    url = api_base_url.rstrip("/") + f"/receipts/{receipt_id}"
    body = api_request(session, "GET", url, timeout_seconds=timeout_seconds)
    if not isinstance(body, dict):
        raise DreamkasError("Некорректный ответ Dreamkas при получении чека")
    return body


def receipt_type_to_fns_n(receipt_type: str) -> str:
    mapping = {
        "SALE": "1",
        "REFUND": "2",
        "OUTFLOW": "3",
        "OUTFLOW_REFUND": "4",
    }
    return mapping.get(receipt_type.upper(), "1")


def parse_receipt_datetime(receipt: dict[str, Any]) -> datetime:
    raw = receipt.get("localDate") or receipt.get("date") or ""
    if not raw:
        return datetime.now()

    text = str(raw).replace("Z", "+00:00")
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S.%f%z", "%Y-%m-%dT%H:%M:%S%z", "%Y-%m-%dT%H:%M:%S.%f"):
        try:
            dt = datetime.strptime(text, fmt)
            return dt.replace(tzinfo=None)
        except ValueError:
            pass
    try:
        return datetime.fromisoformat(text).replace(tzinfo=None)
    except ValueError:
        return datetime.now()


def build_fns_qr_data(receipt: dict[str, Any]) -> str:
    required_fields = ["fnNumber", "fiscalDocumentNumber", "fiscalDocumentSign", "amount"]
    missing = [field for field in required_fields if receipt.get(field) in (None, "")]
    if missing:
        raise ValueError("В ответе Dreamkas нет полей для QR ФНС: " + ", ".join(missing))

    dt = parse_receipt_datetime(receipt)
    t_value = dt.strftime("%Y%m%dT%H%M")
    amount_kopecks = int(receipt["amount"])
    s_value = f"{amount_kopecks / 100:.2f}"
    n_value = receipt_type_to_fns_n(str(receipt.get("type") or "SALE"))

    return (
        f"t={t_value}"
        f"&s={s_value}"
        f"&fn={receipt['fnNumber']}"
        f"&i={receipt['fiscalDocumentNumber']}"
        f"&fp={receipt['fiscalDocumentSign']}"
        f"&n={n_value}"
    )


def save_qr(qr_data: str, png_path: Path) -> None:
    img = qrcode.make(qr_data)
    img.save(png_path)


def save_receipt_txt(
    txt_path: Path,
    draft: ReceiptDraft,
    operation: dict[str, Any],
    receipt: dict[str, Any],
    qr_data: Optional[str],
) -> None:
    lines: list[str] = []
    lines.append("ЧЕК DREAMKAS")
    lines.append("=" * 60)
    lines.append(f"Дата сохранения: {datetime.now().isoformat(timespec='seconds')}")
    lines.append(f"External ID: {draft.external_id}")
    lines.append(f"Operation ID: {operation.get('id', '-')}")
    lines.append(f"Receipt ID: {receipt.get('id', '-')}")
    if draft.parent_external_id:
        lines.append(f"Исходный External ID: {draft.parent_external_id}")
    if draft.parent_receipt_id:
        lines.append(f"Исходный Receipt ID: {draft.parent_receipt_id}")
    if draft.refund_mode:
        lines.append(f"Режим возврата: {draft.refund_mode}")
    if draft.refunded_positions:
        lines.append("Позиции к возврату: " + ", ".join(str(i) for i in draft.refunded_positions))
    lines.append(f"Статус операции: {operation.get('status', '-')}")
    lines.append(f"Тип чека: {receipt.get('type', draft.payload.get('type', '-'))}")
    lines.append(f"Сумма: {format_money(int(receipt.get('amount', draft.total_kopecks)))}")
    lines.append(f"Тип оплаты: {draft.payment_type}")
    lines.append(f"Email: {draft.buyer_email or '-'}")
    lines.append(f"Телефон: {draft.buyer_phone or '-'}")
    lines.append(f"Тип покупателя: {'Юрлицо / ИП' if draft.buyer_type == 'LEGAL_ENTITY' else 'Физлицо'}")
    if draft.buyer_type == 'LEGAL_ENTITY':
        lines.append(f"Покупатель: {draft.buyer_legal_name or '-'}")
        lines.append(f"ИНН покупателя: {draft.buyer_inn or '-'}")
    lines.append(f"СНО: {draft.tax_mode}")
    lines.append(f"Кассир из Excel: {draft.cashier_name}")
    dreamkas_cashier = receipt.get("cashier") if isinstance(receipt, dict) else None
    if isinstance(dreamkas_cashier, dict):
        lines.append(f"Кассир в ответе Dreamkas: {dreamkas_cashier.get('name', '-')}")
    lines.append("")
    lines.append("Фискальные реквизиты:")
    lines.append(f"  ФН:  {receipt.get('fnNumber', '-')}")
    lines.append(f"  ФД:  {receipt.get('fiscalDocumentNumber', '-')}")
    lines.append(f"  ФП:  {receipt.get('fiscalDocumentSign', '-')}")
    lines.append(f"  РН ККТ: {receipt.get('registryNumber', '-')}")
    lines.append(f"  Сайт проверки: {receipt.get('checkURL', '-')}")
    if qr_data:
        lines.append(f"  Данные QR ФНС: {qr_data}")
    lines.append("")
    lines.append("Позиции:")
    for i, p in enumerate(draft.positions, start=1):
        lines.append(
            f"  {i}. {p.name} | {p.item_type} | кол-во {p.quantity} | "
            f"цена {format_money(p.price_kopecks)} | сумма {format_money(p.price_sum_kopecks)} | НДС {p.tax}"
        )
    lines.append("")
    lines.append("JSON операции:")
    lines.append(json.dumps(operation, ensure_ascii=False, indent=2))
    lines.append("")
    lines.append("JSON чека:")
    lines.append(json.dumps(receipt, ensure_ascii=False, indent=2))
    txt_path.write_text("\n".join(lines), encoding="utf-8")



def find_pdf_font() -> tuple[str, Optional[str]]:
    candidates = [
        r"C:\Windows\Fonts\arial.ttf",
        r"C:\Windows\Fonts\Arial.ttf",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "/usr/share/fonts/truetype/liberation2/LiberationSans-Regular.ttf",
        "/System/Library/Fonts/Supplemental/Arial Unicode.ttf",
        "/System/Library/Fonts/Supplemental/Arial.ttf",
    ]
    for item in candidates:
        path = Path(item)
        if path.exists():
            return "ReceiptFont", str(path)
    return "Helvetica", None


def receipt_first_value(obj: Any, *keys: str) -> Optional[Any]:
    """Рекурсивно ищет первое непустое значение по нескольким возможным именам ключей."""
    wanted = {k.lower() for k in keys}

    def walk(value: Any) -> Optional[Any]:
        if isinstance(value, dict):
            for k, v in value.items():
                if str(k).lower() in wanted and v not in (None, ""):
                    return v
            for v in value.values():
                found = walk(v)
                if found not in (None, ""):
                    return found
        elif isinstance(value, list):
            for item in value:
                found = walk(item)
                if found not in (None, ""):
                    return found
        return None

    return walk(obj)


def receipt_str(value: Any, default: str = "-") -> str:
    if value in (None, ""):
        return default
    return str(value)


def receipt_operation_ru(operation_type: str) -> str:
    mapping = {
        "SALE": "ПРИХОД",
        "REFUND": "ВОЗВРАТ ПРИХОДА",
        "SALE_REFUND": "ВОЗВРАТ ПРИХОДА",
        "PURCHASE": "РАСХОД",
        "PURCHASE_REFUND": "ВОЗВРАТ РАСХОДА",
    }
    return mapping.get(str(operation_type or "").upper(), str(operation_type or "-"))


def payment_type_ru(payment_type: str) -> str:
    mapping = {
        "CASH": "НАЛИЧНЫМИ",
        "CASHLESS": "БЕЗНАЛИЧНЫМИ",
        "CARD": "БЕЗНАЛИЧНЫМИ",
        "ELECTRON": "БЕЗНАЛИЧНЫМИ",
    }
    return mapping.get(str(payment_type or "").upper(), str(payment_type or "-"))


def tax_ru(tax: str) -> str:
    mapping = {
        "NDS_NO_TAX": "БЕЗ НДС",
        "NO_VAT": "БЕЗ НДС",
        "NONE": "БЕЗ НДС",
        "VAT_0": "НДС 0%",
        "NDS_0": "НДС 0%",
        "VAT_10": "НДС 10%",
        "NDS_10": "НДС 10%",
        "VAT_20": "НДС 20%",
        "NDS_20": "НДС 20%",
        "VAT_10_110": "НДС 10/110",
        "NDS_10_110": "НДС 10/110",
        "VAT_20_120": "НДС 20/120",
        "NDS_20_120": "НДС 20/120",
    }
    return mapping.get(str(tax or "").upper(), str(tax or "-"))


def item_type_ru(item_type: str) -> str:
    mapping = {
        "SERVICE": "УСЛУГА",
        "COUNTABLE": "ТОВАР",
        "PIECE": "ТОВАР",
        "WEIGHT": "ТОВАР",
        "PRODUCT": "ТОВАР",
    }
    return mapping.get(str(item_type or "").upper(), str(item_type or "-"))


def paragraph_safe(text: Any) -> str:
    import html
    return html.escape(receipt_str(text, ""))


def save_receipt_pdf(
    pdf_path: Path,
    draft: ReceiptDraft,
    operation: dict[str, Any],
    receipt: dict[str, Any],
    qr_path: Optional[Path],
    qr_data: Optional[str],
) -> None:
    """
    Сохраняет PDF в виде кассовой ленты.

    Это не замена фискальному документу из ККТ/ОФД, а удобная печатная/архивная
    копия с основными реквизитами: продавец, дата/место расчета, позиции,
    итоги, форма оплаты, ФН/ФД/ФП, РН ККТ, QR-код и ссылка проверки.
    """
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image, HRFlowable
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    font_name, font_path = find_pdf_font()
    if font_path:
        try:
            pdfmetrics.registerFont(TTFont(font_name, font_path))
        except Exception:
            font_name = "Helvetica"

    # 80 мм — типичная ширина кассовой ленты. Высоту делаем расчетной,
    # чтобы длинный чек помещался на один узкий PDF-лист.
    positions_count = max(1, len(draft.positions))
    estimated_height_mm = max(230, 170 + positions_count * 26)
    page_width = 80 * mm
    page_height = estimated_height_mm * mm

    pdf_path.parent.mkdir(parents=True, exist_ok=True)
    doc = SimpleDocTemplate(
        str(pdf_path),
        pagesize=(page_width, page_height),
        rightMargin=4 * mm,
        leftMargin=4 * mm,
        topMargin=5 * mm,
        bottomMargin=5 * mm,
    )

    styles = getSampleStyleSheet()
    base = ParagraphStyle(
        "ReceiptTapeBase",
        parent=styles["Normal"],
        fontName=font_name,
        fontSize=7.0,
        leading=8.4,
        alignment=1,
        spaceAfter=1.2,
    )
    small = ParagraphStyle(
        "ReceiptTapeSmall",
        parent=base,
        fontSize=6.4,
        leading=7.5,
        alignment=0,
        spaceAfter=0.8,
    )
    center = ParagraphStyle(
        "ReceiptTapeCenter",
        parent=base,
        alignment=1,
        fontSize=7.0,
        leading=8.4,
        spaceAfter=1.2,
    )
    title_style = ParagraphStyle(
        "ReceiptTapeTitle",
        parent=base,
        fontSize=8.4,
        leading=10.0,
        alignment=1,
        spaceAfter=2,
    )
    bold = ParagraphStyle(
        "ReceiptTapeBold",
        parent=base,
        fontSize=7.2,
        leading=8.6,
        alignment=0,
        spaceAfter=1,
    )
    right = ParagraphStyle(
        "ReceiptTapeRight",
        parent=base,
        fontSize=7.0,
        leading=8.4,
        alignment=2,
        spaceAfter=1,
    )

    def p(text: Any, style: ParagraphStyle = small) -> Paragraph:
        return Paragraph(paragraph_safe(text), style)

    def line() -> HRFlowable:
        return HRFlowable(width="100%", thickness=0.35, color=colors.black, spaceBefore=2, spaceAfter=2)

    def kv(label: str, value: Any) -> Table:
        table = Table(
            [[p(label, small), p(value, small)]],
            colWidths=[27 * mm, 45 * mm],
            hAlign="LEFT",
        )
        table.setStyle(TableStyle([
            ("FONT", (0, 0), (-1, -1), font_name, 6.4),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("ALIGN", (1, 0), (1, 0), "RIGHT"),
            ("LEFTPADDING", (0, 0), (-1, -1), 0),
            ("RIGHTPADDING", (0, 0), (-1, -1), 0),
            ("TOPPADDING", (0, 0), (-1, -1), 0),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 0.5),
        ]))
        return table

    op_type = str(receipt.get("type") or draft.payload.get("type") or "").upper()
    op_ru = receipt_operation_ru(op_type)
    receipt_id = receipt.get("id", "-")

    # Реквизиты продавца/места расчета. Dreamkas может вернуть разные имена ключей,
    # поэтому ищем несколько популярных вариантов рекурсивно.
    seller_name = receipt_first_value(
        receipt,
        "user", "userName", "organizationName", "companyName", "legalName", "retailPlace",
        "shopName", "sellerName",
    )
    seller_inn = receipt_first_value(receipt, "inn", "userInn", "companyInn", "sellerInn")
    payment_address = receipt_first_value(
        receipt,
        "address", "paymentAddress", "retailPlaceAddress", "calculationAddress", "realAddress"
    )
    calculation_place = receipt_first_value(receipt, "retailPlace", "paymentPlace", "calculationPlace", "place")

    created_at = (
        receipt.get("createdAt")
        or receipt.get("processedAt")
        or receipt.get("fiscalizedAt")
        or operation.get("createdAt")
        or operation.get("updatedAt")
        or datetime.now().isoformat(timespec="seconds")
    )

    amount = int(receipt.get("amount", draft.total_kopecks) or draft.total_kopecks)

    fn = receipt.get("fnNumber") or receipt_first_value(receipt, "fnNumber", "fiscalDriveNumber", "fn")
    fd = receipt.get("fiscalDocumentNumber") or receipt_first_value(receipt, "fiscalDocumentNumber", "fd")
    fp = receipt.get("fiscalDocumentSign") or receipt_first_value(receipt, "fiscalDocumentSign", "fiscalSign", "fp", "fpd")
    registry_number = receipt.get("registryNumber") or receipt_first_value(receipt, "registryNumber", "kktRegNumber", "kktRegistrationNumber")
    kkt_number = receipt_first_value(receipt, "kktFactoryNumber", "kktSerialNumber", "serialNumber")
    shift_number = receipt_first_value(receipt, "shiftNumber", "shift")
    doc_number = receipt_first_value(receipt, "documentNumber", "receiptNumber", "requestNumber")
    check_url = receipt.get("checkURL") or receipt_first_value(receipt, "checkURL", "ofdCheckUrl", "checkUrl", "url")

    dreamkas_cashier = receipt.get("cashier") if isinstance(receipt, dict) else None
    cashier_from_receipt = None
    if isinstance(dreamkas_cashier, dict):
        cashier_from_receipt = dreamkas_cashier.get("name")
    cashier_name = cashier_from_receipt or draft.cashier_name or "-"

    story: list[Any] = []

    story.append(p("КАССОВЫЙ ЧЕК", title_style))
    story.append(p(op_ru, title_style))
    story.append(line())

    if seller_name:
        story.append(p(str(seller_name).upper(), center))
    if seller_inn:
        story.append(p(f"ИНН {seller_inn}", center))
    if payment_address:
        story.append(p(payment_address, center))
    if calculation_place and calculation_place != payment_address:
        story.append(p(f"МЕСТО РАСЧЕТОВ: {calculation_place}", center))

    story.append(line())
    story.append(kv("ДАТА/ВРЕМЯ", created_at))
    if shift_number:
        story.append(kv("СМЕНА", shift_number))
    if doc_number:
        story.append(kv("№ ЗА СМЕНУ", doc_number))
    story.append(kv("ПРИЗНАК", op_ru))
    story.append(kv("СНО", draft.tax_mode))
    story.append(kv("КАССИР", cashier_name))
    if draft.buyer_email:
        story.append(kv("EMAIL ПОКУП.", draft.buyer_email))
    if draft.buyer_phone:
        story.append(kv("ТЕЛ. ПОКУП.", draft.buyer_phone))
    story.append(kv("ПОКУПАТЕЛЬ", "ЮРЛИЦО / ИП" if draft.buyer_type == "LEGAL_ENTITY" else "ФИЗЛИЦО"))
    if draft.buyer_type == "LEGAL_ENTITY":
        if draft.buyer_legal_name:
            story.append(kv("НАИМ. ПОКУП.", draft.buyer_legal_name))
        if draft.buyer_inn:
            story.append(kv("ИНН ПОКУП.", draft.buyer_inn))

    story.append(line())

    for i, pos in enumerate(draft.positions, start=1):
        story.append(p(f"{i}. {pos.name}", bold))
        story.append(kv("ПРЕДМЕТ", item_type_ru(pos.item_type)))
        story.append(kv("КОЛ-ВО", f"{pos.quantity} x {format_money(pos.price_kopecks)}"))
        story.append(kv("СУММА", format_money(pos.price_sum_kopecks)))
        story.append(kv("НДС", tax_ru(pos.tax)))
        if i != len(draft.positions):
            story.append(Spacer(1, 1.5))

    story.append(line())

    # Итоги
    story.append(kv("ИТОГ", format_money(amount)))
    story.append(kv(payment_type_ru(draft.payment_type), format_money(amount)))

    # НДС-итоги по ставкам
    vat_groups: dict[str, int] = {}
    for pos in draft.positions:
        vat_groups[tax_ru(pos.tax)] = vat_groups.get(tax_ru(pos.tax), 0) + int(pos.price_sum_kopecks)
    for tax_name, tax_base in vat_groups.items():
        if tax_name == "БЕЗ НДС":
            story.append(kv("ИТОГ БЕЗ НДС", format_money(tax_base)))
        else:
            story.append(kv(tax_name, format_money(tax_base)))

    story.append(line())

    if registry_number:
        story.append(kv("РН ККТ", registry_number))
    if kkt_number:
        story.append(kv("ЗН ККТ", kkt_number))
    if fn:
        story.append(kv("ФН", fn))
    if fd:
        story.append(kv("ФД", fd))
    if fp:
        story.append(kv("ФП", fp))

    story.append(kv("RECEIPT ID", receipt_id))
    story.append(kv("OPERATION ID", operation.get("id", "-")))
    story.append(kv("EXTERNAL ID", draft.external_id))

    if draft.parent_external_id:
        story.append(kv("ИСХ. EXTERNAL", draft.parent_external_id))
    if draft.parent_receipt_id:
        story.append(kv("ИСХ. RECEIPT", draft.parent_receipt_id))
    if draft.refund_mode:
        story.append(kv("РЕЖИМ ВОЗВР.", draft.refund_mode))

    if check_url:
        story.append(line())
        story.append(p("АДРЕС ПРОВЕРКИ ЧЕКА", center))
        story.append(p(check_url, small))

    if qr_path and qr_path.exists():
        story.append(line())
        story.append(p("QR-КОД ДЛЯ ПРОВЕРКИ", center))
        # Минимум по современным требованиям — 20x20 мм, поэтому используем 25x25 мм.
        img = Image(str(qr_path), width=25 * mm, height=25 * mm)
        img.hAlign = "CENTER"
        story.append(img)

    if qr_data:
        story.append(Spacer(1, 2))
        story.append(p(qr_data, small))

    story.append(line())
    story.append(p("СПАСИБО ЗА ПОКУПКУ!", center))
    story.append(p("Электронная копия сформирована Dreamkas Receipt Tool", center))

    doc.build(story)



def send_receipt_email(
    settings_path: Path,
    settings: dict[str, str],
    draft: ReceiptDraft,
    receipt: dict[str, Any],
    txt_path: Optional[Path],
    pdf_path: Optional[Path],
    qr_path: Optional[Path],
) -> bool:
    if not draft.buyer_email:
        return False
    if not bool_setting(settings, "EMAIL_ENABLED_AFTER_SUCCESS", False):
        return False

    smtp_host = settings.get("SMTP_HOST", "").strip()
    smtp_from = settings.get("SMTP_FROM", "").strip() or settings.get("SMTP_USER", "").strip()
    if not smtp_host or not smtp_from:
        print("\nEmail не отправлен: SMTP_HOST/SMTP_FROM не настроены в settings.txt")
        return False

    answer = input(f"\nОтправить чек покупателю на email {draft.buyer_email}? Введите ДА: ").strip().lower()
    if answer not in {"да", "д", "yes", "y"}:
        return False

    msg = EmailMessage()
    msg["Subject"] = f"Ваш кассовый чек {receipt.get('id', '')}"
    msg["From"] = smtp_from
    msg["To"] = draft.buyer_email
    check_url = receipt.get("checkURL") or ""
    msg.set_content(
        "Здравствуйте!\n\n"
        "Во вложении кассовый чек.\n"
        + (f"Ссылка на проверку чека: {check_url}\n" if check_url else "")
        + "\nЭто письмо сформировано автоматически.\n"
    )

    for path in [pdf_path, txt_path, qr_path]:
        if path and path.exists():
            data = path.read_bytes()
            maintype = "application"
            subtype = "octet-stream"
            if path.suffix.lower() == ".pdf":
                subtype = "pdf"
            elif path.suffix.lower() == ".txt":
                maintype, subtype = "text", "plain"
            elif path.suffix.lower() == ".png":
                maintype, subtype = "image", "png"
            msg.add_attachment(data, maintype=maintype, subtype=subtype, filename=path.name)

    port = get_int(settings, "SMTP_PORT", 587)
    use_ssl = bool_setting(settings, "SMTP_SSL", False)
    use_starttls = bool_setting(settings, "SMTP_STARTTLS", True)
    user = settings.get("SMTP_USER", "").strip()
    password = get_smtp_password(settings_path, settings, ask_if_missing=True)
    context = smtp_ssl_context(settings)

    try:
        if use_ssl:
            server = smtplib.SMTP_SSL(smtp_host, port, timeout=30, context=context)
        else:
            server = smtplib.SMTP(smtp_host, port, timeout=30)
        with server:
            if use_starttls and not use_ssl:
                server.starttls(context=context)
            if user:
                server.login(user, password)
            server.send_message(msg)
        print("Email с чеком отправлен.")
        logging.info("Receipt email sent to %s", draft.buyer_email)
        return True
    except Exception as exc:
        print(f"Email не отправлен: {exc}")
        logging.exception("Could not send receipt email")
        return False



def enable_ansi_colors() -> None:
    """Включает ANSI-цвета в современном Windows Terminal/CMD, где это возможно."""
    if os.name == "nt":
        try:
            os.system("")
        except Exception:
            pass


def selected_line(text: str) -> str:
    """Выделение выбранных к возврату позиций. Если цвет не сработает, остается текстовая метка."""
    return f"\033[30;47m{text}\033[0m  <<< К ВОЗВРАТУ"


def load_json_object(raw: Any) -> dict[str, Any]:
    if not raw:
        return {}
    if isinstance(raw, dict):
        return raw
    try:
        obj = json.loads(str(raw))
    except Exception:
        return {}
    return obj if isinstance(obj, dict) else {}


def payload_positions(payload: dict[str, Any]) -> list[dict[str, Any]]:
    positions = payload.get("positions") or []
    return [p for p in positions if isinstance(p, dict)]


def position_sum_kopecks(pos: dict[str, Any]) -> int:
    if pos.get("priceSum") not in (None, ""):
        return int(Decimal(str(pos.get("priceSum"))).quantize(Decimal("1"), rounding=ROUND_HALF_UP))
    price = Decimal(str(pos.get("price") or "0"))
    quantity = Decimal(str(pos.get("quantity") or "1"))
    return int((price * quantity).quantize(Decimal("1"), rounding=ROUND_HALF_UP))


def position_to_receipt_position(pos: dict[str, Any]) -> ReceiptPosition:
    quantity = Decimal(str(pos.get("quantity") or "1"))
    price = int(Decimal(str(pos.get("price") or "0")).quantize(Decimal("1"), rounding=ROUND_HALF_UP))
    return ReceiptPosition(
        name=str(pos.get("name") or "Без названия"),
        item_type=str(pos.get("type") or "COUNTABLE"),
        quantity=quantity,
        price_kopecks=price,
        price_sum_kopecks=position_sum_kopecks(pos),
        tax=str(pos.get("tax") or "NDS_NO_TAX"),
    )


def first_payment_type(payload: dict[str, Any]) -> str:
    payments = payload.get("payments") or []
    if isinstance(payments, list) and payments:
        first = payments[0]
        if isinstance(first, dict) and first.get("type"):
            return str(first.get("type"))
    return "CASHLESS"



def legal_buyer_from_payload(payload: dict[str, Any]) -> tuple[str, Optional[str], Optional[str]]:
    """Пробует извлечь данные юрлица из ФФД-тегов 1227/1228 в payload."""
    legal_name: Optional[str] = None
    buyer_inn: Optional[str] = None
    tags = payload.get("tags")
    if isinstance(tags, list):
        for tag in tags:
            if not isinstance(tag, dict):
                continue
            if tag.get("tag") == 1227:
                legal_name = str(tag.get("value") or "").strip() or legal_name
            elif tag.get("tag") == 1228:
                buyer_inn = normalize_buyer_inn(tag.get("value")) or buyer_inn
    return ("LEGAL_ENTITY" if legal_name or buyer_inn else "INDIVIDUAL", legal_name, buyer_inn)

def buyer_from_payload(payload: dict[str, Any]) -> tuple[Optional[str], Optional[str]]:
    attributes = payload.get("attributes") or {}
    if not isinstance(attributes, dict):
        return None, None
    email = str(attributes.get("email") or "").strip() or None
    phone = str(attributes.get("phone") or "").strip() or None
    return email, phone


def row_receipt_type(row: sqlite3.Row) -> str:
    stored = str(row["receipt_type"] or "").upper() if "receipt_type" in row.keys() else ""
    if stored:
        return stored
    return payload_type_from_json(row["request_json"])


def successful_sale_rows(conn: sqlite3.Connection) -> list[sqlite3.Row]:
    rows = conn.execute(
        """
        SELECT * FROM receipts
        WHERE UPPER(status) = 'SUCCESS'
        ORDER BY COALESCE(completed_at, sent_at, created_at) DESC, id DESC
        """
    ).fetchall()
    result: list[sqlite3.Row] = []
    for row in rows:
        if row_receipt_type(row) != "SALE":
            continue
        payload = load_json_object(row["request_json"])
        if not payload_positions(payload):
            continue
        result.append(row)
    return result


def refunded_amounts_by_parent(conn: sqlite3.Connection) -> dict[str, int]:
    rows = conn.execute(
        """
        SELECT parent_external_id, SUM(COALESCE(amount_kopecks, 0)) AS amount
        FROM receipts
        WHERE UPPER(status) = 'SUCCESS'
          AND parent_external_id IS NOT NULL
          AND parent_external_id != ''
        GROUP BY parent_external_id
        """
    ).fetchall()
    return {str(row["parent_external_id"]): int(row["amount"] or 0) for row in rows}


def choose_work_mode() -> str:
    print("\nЧто будем делать?")
    print("  1. Фискализация нового чека")
    print("  2. Возврат ранее пробитого чека")
    print("  3. История чеков")
    print("  4. Настройки")
    print("  0. Выход")
    while True:
        answer = input("Выберите режим: ").strip().lower()
        if answer in {"1", "ф", "фискализация"}:
            return "sale"
        if answer in {"2", "в", "возврат"}:
            return "refund"
        if answer in {"3", "и", "история", "history"}:
            return "history"
        if answer in {"4", "н", "настройки", "settings"}:
            return "settings"
        if answer in {"0", "q", "й", "exit", "выход"}:
            return "exit"
        print("Некорректный выбор. Введите 1, 2, 3, 4 или 0.")


def receipt_rows(conn: sqlite3.Connection, limit: int = 50, only_today: bool = False, search: str = "") -> list[sqlite3.Row]:
    query = "SELECT * FROM receipts WHERE 1=1"
    params: list[Any] = []
    if only_today:
        query += " AND substr(created_at, 1, 10) = ?"
        params.append(datetime.now().strftime("%Y-%m-%d"))
    if search:
        like = f"%{search}%"
        query += " AND (buyer_email LIKE ? OR buyer_phone LIKE ? OR receipt_id LIKE ? OR external_id LIKE ? OR cashier_name LIKE ?)"
        params.extend([like, like, like, like, like])
    query += " ORDER BY id DESC LIMIT ?"
    params.append(limit)
    return conn.execute(query, params).fetchall()


def print_receipt_rows(rows: list[sqlite3.Row], title: str = "История чеков") -> None:
    print(f"\n{title}")
    print("-" * 118)
    print(f"{'№':>3} {'DB':>5} {'Дата':<19} {'Тип':<8} {'Статус':<11} {'Сумма':>12} {'Кассир':<18} {'Покупатель':<24} {'Receipt ID'}")
    print("-" * 118)
    for i, row in enumerate(rows, start=1):
        buyer = row["buyer_email"] or row["buyer_phone"] or "-"
        dt = str(row["completed_at"] or row["sent_at"] or row["created_at"] or "-")[:19]
        receipt_type = row_receipt_type(row) or "-"
        print(
            f"{i:>3} {row['id']:>5} {dt:<19} {receipt_type:<8} {str(row['status'] or '-'):<11} "
            f"{format_money(int(row['amount_kopecks'] or 0)):>12} {str(row['cashier_name'] or '-')[:18]:<18} "
            f"{str(buyer)[:24]:<24} {row['receipt_id'] or '-'}"
        )
    print("-" * 118)


def show_history_menu(conn: sqlite3.Connection) -> None:
    while True:
        print("\nИстория чеков:")
        print("  1. Последние 20")
        print("  2. Только за сегодня")
        print("  3. Поиск по телефону/email/receiptId")
        print("  0. Назад")
        answer = input("Выберите вариант: ").strip()
        if answer in {"0", "", "q", "й"}:
            return
        if answer == "1":
            print_receipt_rows(receipt_rows(conn, limit=20), "Последние 20 чеков")
            wait_key("Нажмите любую клавишу для возврата в меню...")
            continue
        if answer == "2":
            print_receipt_rows(receipt_rows(conn, limit=100, only_today=True), "Чеки за сегодня")
            wait_key("Нажмите любую клавишу для возврата в меню...")
            continue
        if answer == "3":
            q = input("Введите строку поиска: ").strip()
            print_receipt_rows(receipt_rows(conn, limit=100, search=q), f"Поиск: {q}")
            wait_key("Нажмите любую клавишу для возврата в меню...")
            continue
        print("Некорректный выбор.")


def settings_menu(settings_path: Path, settings: dict[str, str], session: requests.Session, api_base_url: str, request_timeout: int) -> dict[str, str]:
    while True:
        print("\nНастройки:")
        print("  1. Изменить токен Dreamkas")
        print("  2. Заново выбрать магазин и кассу")
        print("  3. Изменить интервал проверки статуса")
        print("  4. Изменить окно защиты от дублей")
        print("  5. Настроить SMTP для отправки email")
        print("  6. Включить/выключить PDF")
        print("  7. Включить/выключить очистку Excel после успешного чека")
        print("  8. Настроить поиск юрлица / ИП по ИНН")
        print("  9. Выбрать способ безналичной оплаты")
        print(" 10. Настроить СБП через T-Банк")
        print(" 11. Настроить СБП B2B/I2I через T-Банк")
        print("  0. Назад")
        answer = input("Выберите пункт: ").strip()
        if answer in {"0", "", "q", "й"}:
            save_settings(settings_path, settings)
            return settings
        if answer == "1":
            settings["DREAMKAS_TOKEN"] = ask_token_from_user()
            session.headers.update(dreamkas_headers(settings["DREAMKAS_TOKEN"]))
            settings = clear_smtp_settings(settings_path, settings, "DREAMKAS_TOKEN изменен пользователем")
            settings["TBANK_PASSWORD_ENC"] = ""
            settings["TBANK_B2B_API_TOKEN_ENC"] = ""
            save_settings(settings_path, settings)
            print("Токен обновлен. SMTP-пароль, пароль T-Банка и T-API token очищены, потому что ключ шифрования изменился.")
            continue
        if answer == "2":
            settings = select_shop_and_device(settings_path, settings, session, api_base_url, request_timeout)
            continue
        if answer == "3":
            value = input(f"POLL_INTERVAL_SECONDS [{settings.get('POLL_INTERVAL_SECONDS', '20')}]: ").strip()
            if value:
                settings["POLL_INTERVAL_SECONDS"] = value
            save_settings(settings_path, settings)
            continue
        if answer == "4":
            print("Защита от дублей теперь работает по полному предчеку из SQLite.")
            value = input(f"DUPLICATE_LOOKBACK_DAYS [{settings.get('DUPLICATE_LOOKBACK_DAYS', '30')}]: ").strip()
            if value:
                settings["DUPLICATE_LOOKBACK_DAYS"] = value
            save_settings(settings_path, settings)
            continue
        if answer == "5":
            settings["EMAIL_ENABLED_AFTER_SUCCESS"] = input("Отправлять email после успешного чека? 1/0: ").strip() or settings.get("EMAIL_ENABLED_AFTER_SUCCESS", "0")
            settings["SMTP_HOST"] = input(f"SMTP_HOST [{settings.get('SMTP_HOST','')}]: ").strip() or settings.get("SMTP_HOST", "")
            settings["SMTP_PORT"] = input(f"SMTP_PORT [{settings.get('SMTP_PORT','587')}]: ").strip() or settings.get("SMTP_PORT", "587")
            settings["SMTP_USER"] = input(f"SMTP_USER [{settings.get('SMTP_USER','')}]: ").strip() or settings.get("SMTP_USER", "")
            if input("Изменить SMTP_PASSWORD? 1/0: ").strip() == "1":
                new_password = getpass.getpass("SMTP_PASSWORD: ").strip()
                if new_password:
                    settings["SMTP_PASSWORD_ENC"] = encrypt_smtp_password(new_password, settings.get("DREAMKAS_TOKEN", ""))
                    settings.pop("SMTP_PASSWORD", None)
                    print("SMTP-пароль зашифрован и сохранен как SMTP_PASSWORD_ENC.")
            settings["SMTP_FROM"] = input(f"SMTP_FROM [{settings.get('SMTP_FROM','')}]: ").strip() or settings.get("SMTP_FROM", "")
            settings["SMTP_SSL"] = input(f"SMTP_SSL 1/0 [{settings.get('SMTP_SSL','0')}]: ").strip() or settings.get("SMTP_SSL", "0")
            settings["SMTP_STARTTLS"] = input(f"SMTP_STARTTLS 1/0 [{settings.get('SMTP_STARTTLS','1')}]: ").strip() or settings.get("SMTP_STARTTLS", "1")
            settings["SMTP_VERIFY_CERT"] = input(f"SMTP_VERIFY_CERT 1/0 [{settings.get('SMTP_VERIFY_CERT','1')}]: ").strip() or settings.get("SMTP_VERIFY_CERT", "1")
            save_settings(settings_path, settings)
            continue
        if answer == "6":
            settings["PDF_ENABLED"] = "0" if bool_setting(settings, "PDF_ENABLED", True) else "1"
            save_settings(settings_path, settings)
            print(f"PDF_ENABLED = {settings['PDF_ENABLED']}")
            continue
        if answer == "7":
            settings["CLEAR_EXCEL_AFTER_SUCCESS"] = "0" if bool_setting(settings, "CLEAR_EXCEL_AFTER_SUCCESS", True) else "1"
            save_settings(settings_path, settings)
            print(f"CLEAR_EXCEL_AFTER_SUCCESS = {settings['CLEAR_EXCEL_AFTER_SUCCESS']}")
            continue
        if answer == "8":
            print("\nПоиск названия юрлица / ИП по ИНН")
            settings["LEGAL_ENTITY_LOOKUP_ENABLED"] = input(
                f"Включить поиск по ИНН? 1/0 [{settings.get('LEGAL_ENTITY_LOOKUP_ENABLED','1')}]: "
            ).strip() or settings.get("LEGAL_ENTITY_LOOKUP_ENABLED", "1")
            settings["LEGAL_ENTITY_LOOKUP_PROVIDER"] = input(
                f"Провайдер [DADATA] [{settings.get('LEGAL_ENTITY_LOOKUP_PROVIDER','DADATA')}]: "
            ).strip() or settings.get("LEGAL_ENTITY_LOOKUP_PROVIDER", "DADATA")
            print("Для DaData нужен API token из личного кабинета DaData.")
            if input("Изменить DADATA_TOKEN? 1/0: ").strip() == "1":
                try:
                    token = getpass.getpass("DADATA_TOKEN: ").strip()
                except Exception:
                    token = input("DADATA_TOKEN: ").strip()
                if token:
                    settings["DADATA_TOKEN"] = token
            settings["DADATA_BRANCH_TYPE"] = input(
                f"DADATA_BRANCH_TYPE MAIN/BRANCH [{settings.get('DADATA_BRANCH_TYPE','MAIN')}]: "
            ).strip() or settings.get("DADATA_BRANCH_TYPE", "MAIN")
            settings["DADATA_UPDATE_EXCEL_NAME"] = input(
                f"Записывать найденное название обратно в Excel? 1/0 [{settings.get('DADATA_UPDATE_EXCEL_NAME','1')}]: "
            ).strip() or settings.get("DADATA_UPDATE_EXCEL_NAME", "1")
            save_settings(settings_path, settings)
            continue
        if answer == "9":
            print("\nСпособ безналичной оплаты")
            print("  1. Внешний банковский терминал")
            print("  2. T-Банк СБП")
            current_provider = normalize_cashless_payment_provider(settings.get("CASHLESS_PAYMENT_PROVIDER", "EXTERNAL_TERMINAL"))
            print(f"Текущий способ: {cashless_provider_label(current_provider)}")
            provider_answer = input("Выберите способ [1/2]: ").strip()
            if provider_answer == "1":
                settings["CASHLESS_PAYMENT_PROVIDER"] = "EXTERNAL_TERMINAL"
                settings["TBANK_SBP_ENABLED"] = "0"
            elif provider_answer == "2":
                settings["CASHLESS_PAYMENT_PROVIDER"] = "TBANK_SBP"
                settings["TBANK_SBP_ENABLED"] = "1"
            else:
                print("Способ не изменён.")
            save_settings(settings_path, settings)
            print(f"CASHLESS_PAYMENT_PROVIDER = {settings.get('CASHLESS_PAYMENT_PROVIDER')}")
            continue
        if answer == "10":
            print("\nНастройка СБП через T-Банк")
            settings["TBANK_SBP_ENABLED"] = input(
                f"Включить оплату СБП через T-Банк? 1/0 [{settings.get('TBANK_SBP_ENABLED','0')}]: "
            ).strip() or settings.get("TBANK_SBP_ENABLED", "0")
            settings["TBANK_TERMINAL_KEY"] = input(
                f"TBANK_TERMINAL_KEY [{settings.get('TBANK_TERMINAL_KEY','')}]: "
            ).strip() or settings.get("TBANK_TERMINAL_KEY", "")
            if input("Изменить TBANK_PASSWORD? 1/0: ").strip() == "1":
                password = getpass.getpass("TBANK_PASSWORD: ").strip()
                if password:
                    settings["TBANK_PASSWORD_ENC"] = encrypt_smtp_password(password, settings.get("DREAMKAS_TOKEN", ""))
                    settings.pop("TBANK_PASSWORD", None)
                    print("Пароль T-Банка зашифрован и сохранен как TBANK_PASSWORD_ENC.")
            settings["TBANK_PAY_TYPE"] = input(
                f"TBANK_PAY_TYPE O/T [{settings.get('TBANK_PAY_TYPE','O')}]: "
            ).strip() or settings.get("TBANK_PAY_TYPE", "O")
            settings["TBANK_PAYMENT_TIMEOUT_MINUTES"] = input(
                f"TBANK_PAYMENT_TIMEOUT_MINUTES [{settings.get('TBANK_PAYMENT_TIMEOUT_MINUTES','10')}]: "
            ).strip() or settings.get("TBANK_PAYMENT_TIMEOUT_MINUTES", "10")
            settings["TBANK_POLL_INTERVAL_SECONDS"] = input(
                f"TBANK_POLL_INTERVAL_SECONDS [{settings.get('TBANK_POLL_INTERVAL_SECONDS','5')}]: "
            ).strip() or settings.get("TBANK_POLL_INTERVAL_SECONDS", "5")
            settings["TBANK_REQUEST_TIMEOUT_SECONDS"] = input(
                f"TBANK_REQUEST_TIMEOUT_SECONDS [{settings.get('TBANK_REQUEST_TIMEOUT_SECONDS','30')}]: "
            ).strip() or settings.get("TBANK_REQUEST_TIMEOUT_SECONDS", "30")
            settings["TBANK_QR_DATA_TYPE"] = input(
                f"TBANK_QR_DATA_TYPE PAYLOAD/IMAGE [{settings.get('TBANK_QR_DATA_TYPE','PAYLOAD')}]: "
            ).strip() or settings.get("TBANK_QR_DATA_TYPE", "PAYLOAD")
            settings["TBANK_DESCRIPTION_PREFIX"] = input(
                f"TBANK_DESCRIPTION_PREFIX [{settings.get('TBANK_DESCRIPTION_PREFIX','Оплата заказа')}]: "
            ).strip() or settings.get("TBANK_DESCRIPTION_PREFIX", "Оплата заказа")

            if bool_setting(settings, "TBANK_SBP_ENABLED", False):
                settings["CASHLESS_PAYMENT_PROVIDER"] = "TBANK_SBP"

            save_settings(settings_path, settings)
            continue
        if answer == "11":
            print("\nНастройка СБП B2B/I2I через T-Банк T-API")
            print("Этот режим используется для оплаты от юрлиц и ИП.")
            settings["TBANK_B2B_SBP_ENABLED"] = input(
                f"Включить СБП B2B/I2I? 1/0 [{settings.get('TBANK_B2B_SBP_ENABLED','0')}]: "
            ).strip() or settings.get("TBANK_B2B_SBP_ENABLED", "0")
            settings["TBANK_B2B_ACCOUNT_NUMBER"] = input(
                f"TBANK_B2B_ACCOUNT_NUMBER, расчетный счет 20/22 цифры [{settings.get('TBANK_B2B_ACCOUNT_NUMBER','')}]: "
            ).strip() or settings.get("TBANK_B2B_ACCOUNT_NUMBER", "")
            if input("Изменить TBANK_B2B_API_TOKEN? 1/0: ").strip() == "1":
                token = getpass.getpass("TBANK_B2B_API_TOKEN: ").strip()
                if token:
                    settings["TBANK_B2B_API_TOKEN_ENC"] = encrypt_smtp_password(token, settings.get("DREAMKAS_TOKEN", ""))
                    settings.pop("TBANK_B2B_API_TOKEN", None)
                    print("T-API token зашифрован и сохранен как TBANK_B2B_API_TOKEN_ENC.")
            settings["TBANK_B2B_TTL_DAYS"] = input(
                f"TBANK_B2B_TTL_DAYS [{settings.get('TBANK_B2B_TTL_DAYS','1')}]: "
            ).strip() or settings.get("TBANK_B2B_TTL_DAYS", "1")
            settings["TBANK_B2B_VAT"] = input(
                f"TBANK_B2B_VAT 0/5/7/10/20/22 [{settings.get('TBANK_B2B_VAT','0')}]: "
            ).strip() or settings.get("TBANK_B2B_VAT", "0")
            settings["TBANK_B2B_REDIRECT_URL"] = input(
                f"TBANK_B2B_REDIRECT_URL [{settings.get('TBANK_B2B_REDIRECT_URL','')}]: "
            ).strip() or settings.get("TBANK_B2B_REDIRECT_URL", "")
            settings["TBANK_B2B_PURPOSE_PREFIX"] = input(
                f"TBANK_B2B_PURPOSE_PREFIX [{settings.get('TBANK_B2B_PURPOSE_PREFIX','Оплата по счету')}]: "
            ).strip() or settings.get("TBANK_B2B_PURPOSE_PREFIX", "Оплата по счету")
            save_settings(settings_path, settings)
            continue
        print("Некорректный выбор.")


def choose_receipt_for_refund(conn: sqlite3.Connection) -> Optional[sqlite3.Row]:
    rows_all = successful_sale_rows(conn)
    if not rows_all:
        print("\nВ SQLite нет успешных чеков продажи, доступных для возврата.")
        return None

    while True:
        print("\nВыбор чека для возврата:")
        print("  1. Последние 20 успешных продаж")
        print("  2. Успешные продажи за сегодня")
        print("  3. Поиск по телефону/email/receiptId")
        print("  0. Отмена")
        mode = input("Выберите вариант: ").strip()
        if mode in {"0", "", "q", "й"}:
            return None
        rows = rows_all
        if mode == "1":
            rows = rows_all[:20]
        elif mode == "2":
            today = datetime.now().strftime("%Y-%m-%d")
            rows = [r for r in rows_all if str(r["completed_at"] or r["sent_at"] or r["created_at"] or "").startswith(today)]
        elif mode == "3":
            q = input("Введите строку поиска: ").strip().lower()
            rows = [
                r for r in rows_all
                if q in str(r["buyer_email"] or "").lower()
                or q in str(r["buyer_phone"] or "").lower()
                or q in str(r["receipt_id"] or "").lower()
                or q in str(r["external_id"] or "").lower()
            ]
        else:
            print("Некорректный выбор.")
            continue

        if not rows:
            print("Подходящих чеков не найдено.")
            continue

        refunded = refunded_amounts_by_parent(conn)
        print("\nУспешные чеки продажи из SQLite:")
        print("-" * 110)
        print(f"{'№':>3} {'DB ID':>6} {'Дата':<20} {'Сумма':>12} {'Возвраты':>12} {'Кассир':<18} {'Покупатель':<22} {'Receipt ID'}")
        print("-" * 110)
        for idx, row in enumerate(rows, start=1):
            buyer = row["buyer_email"] or row["buyer_phone"] or "-"
            date_value = row["completed_at"] or row["sent_at"] or row["created_at"] or "-"
            amount = int(row["amount_kopecks"] or 0)
            refunded_amount = refunded.get(str(row["external_id"]), 0)
            refund_text = format_money(refunded_amount) if refunded_amount else "-"
            print(
                f"{idx:>3} {row['id']:>6} {str(date_value)[:19]:<20} "
                f"{format_money(amount):>12} {refund_text:>12} {str(row['cashier_name'] or '-')[:18]:<18} "
                f"{str(buyer)[:22]:<22} {row['receipt_id'] or '-'}"
            )
        print("-" * 110)
        print("0. Назад")

        while True:
            answer = input("Выберите чек для возврата: ").strip()
            if answer in {"0", "", "q", "й"}:
                break
            if answer.isdigit():
                number = int(answer)
                if 1 <= number <= len(rows):
                    return rows[number - 1]
            print("Некорректный выбор. Введите номер из списка.")


def print_saved_receipt_details(row: sqlite3.Row, selected_indices: Optional[list[int]] = None) -> None:
    selected = set(selected_indices or [])
    payload = load_json_object(row["request_json"])
    receipt = load_json_object(row["receipt_json"])
    positions = payload_positions(payload)

    print("\n" + "=" * 100)
    print("ИНФОРМАЦИЯ ПО ВЫБРАННОМУ ЧЕКУ")
    print("=" * 100)
    print(f"DB ID:               {row['id']}")
    print(f"External ID:         {row['external_id']}")
    print(f"Operation ID:        {row['operation_id'] or '-'}")
    print(f"Receipt ID:          {row['receipt_id'] or '-'}")
    print(f"Дата создания:       {row['created_at'] or '-'}")
    print(f"Дата фискализации:   {row['completed_at'] or '-'}")
    print(f"Тип:                 {row_receipt_type(row) or payload.get('type') or '-'}")
    print(f"Сумма:               {format_money(int(row['amount_kopecks'] or 0))}")
    print(f"Тип оплаты:          {row['payment_type'] or first_payment_type(payload)}")
    print(f"Email:               {row['buyer_email'] or '-'}")
    print(f"Телефон:             {row['buyer_phone'] or '-'}")
    print(f"СНО:                 {row['tax_mode'] or payload.get('taxMode') or '-'}")
    print(f"Кассир:              {row['cashier_name'] or '-'}")
    if receipt:
        print("\nФискальные реквизиты:")
        print(f"  ФН:                {receipt.get('fnNumber', '-')}")
        print(f"  ФД:                {receipt.get('fiscalDocumentNumber', '-')}")
        print(f"  ФП:                {receipt.get('fiscalDocumentSign', '-')}")
        print(f"  РН ККТ:            {receipt.get('registryNumber', '-')}")
        print(f"  Ссылка проверки:   {receipt.get('checkURL', '-')}")

    print("\nПозиции:")
    print("-" * 100)
    header = f"{'№':>2}  {'Наименование':<42} {'Тип':<10} {'Кол-во':>8} {'Цена':>12} {'Сумма':>12} {'НДС':>12}"
    print(header)
    print("-" * 100)
    for idx, pos in enumerate(positions, start=1):
        rp = position_to_receipt_position(pos)
        line = (
            f"{idx:>2}  {rp.name[:42]:<42} {rp.item_type:<10} {str(rp.quantity.normalize()):>8} "
            f"{format_money(rp.price_kopecks):>12} {format_money(rp.price_sum_kopecks):>12} {rp.tax:>12}"
        )
        if idx in selected:
            print(selected_line(line))
        else:
            print(line)
    print("-" * 100)
    if selected:
        total_selected = sum(position_sum_kopecks(positions[i - 1]) for i in selected)
        print(f"ИТОГО К ВОЗВРАТУ: {format_money(total_selected)}")
    print("=" * 100)


def choose_refund_variant(positions_count: int) -> str:
    while True:
        print("\nЧто возвращаем?")
        print("  1. Весь чек")
        if positions_count > 1:
            print("  2. Только часть позиций")
        print("  0. Отмена")
        answer = input("Выберите вариант: ").strip()
        if answer == "1":
            return "full"
        if answer == "2" and positions_count > 1:
            return "partial"
        if answer in {"0", "", "q", "й"}:
            return "cancel"
        if answer == "2" and positions_count == 1:
            print("В этом чеке всего одна позиция. Доступен только возврат всего чека или отмена.")
        else:
            print("Некорректный выбор.")


def parse_position_selection(raw: str, max_number: int) -> list[int]:
    parts = [p.strip() for p in raw.split(",") if p.strip()]
    if not parts:
        raise ValueError("Не выбрана ни одна позиция")
    result: list[int] = []
    for part in parts:
        if not part.isdigit():
            raise ValueError(f"'{part}' не является номером позиции")
        number = int(part)
        if number < 1 or number > max_number:
            raise ValueError(f"Позиции №{number} нет в чеке")
        if number not in result:
            result.append(number)
    return result


def choose_partial_positions(row: sqlite3.Row, positions_count: int) -> Optional[list[int]]:
    while True:
        raw = input("Введите номера позиций к возврату через запятую, например 1,5. 0 = назад: ").strip()
        if raw in {"0", "", "q", "й"}:
            return None
        try:
            selected = parse_position_selection(raw, positions_count)
        except ValueError as exc:
            print(f"Ошибка выбора: {exc}")
            continue
        print_saved_receipt_details(row, selected)
        answer = input("Позиции выбраны верно? Введите ДА для подтверждения: ").strip().lower()
        if answer in {"да", "д", "yes", "y"}:
            return selected
        print("Возвращаюсь к выбору позиций.")


def build_refund_draft_from_row(
    row: sqlite3.Row,
    settings: dict[str, str],
    selected_indices: list[int],
    refund_mode: str,
    cashier_override: Optional[str] = None,
) -> ReceiptDraft:
    original_payload = load_json_object(row["request_json"])
    original_positions = payload_positions(original_payload)
    if not original_positions:
        raise ValueError("В сохраненном чеке нет позиций в request_json")

    selected_positions = [deepcopy(original_positions[i - 1]) for i in selected_indices]
    total_kopecks = sum(position_sum_kopecks(pos) for pos in selected_positions)
    if total_kopecks <= 0:
        raise ValueError("Сумма возврата должна быть больше 0")

    buyer_email = row["buyer_email"] or None
    buyer_phone = row["buyer_phone"] or None
    if not buyer_email and not buyer_phone:
        buyer_email, buyer_phone = buyer_from_payload(original_payload)

    precheck = load_json_object(row["precheck_json"] if "precheck_json" in row.keys() else None)
    buyer_type = str(precheck.get("buyer_type") or "") if isinstance(precheck, dict) else ""
    buyer_legal_name = str(precheck.get("buyer_legal_name") or "").strip() if isinstance(precheck, dict) else ""
    buyer_inn = normalize_buyer_inn(precheck.get("buyer_inn")) if isinstance(precheck, dict) else None
    if not buyer_type or buyer_type == "None":
        buyer_type, tag_legal_name, tag_buyer_inn = legal_buyer_from_payload(original_payload)
        buyer_legal_name = buyer_legal_name or (tag_legal_name or "")
        buyer_inn = buyer_inn or tag_buyer_inn
    if not buyer_legal_name:
        buyer_legal_name = None

    attributes: dict[str, str] = {}
    if buyer_email:
        attributes["email"] = str(buyer_email)
    if buyer_phone:
        attributes["phone"] = str(buyer_phone)
    if not attributes:
        attrs = original_payload.get("attributes") or {}
        if isinstance(attrs, dict):
            attributes = dict(attrs)

    payment_type = row["payment_type"] or first_payment_type(original_payload)
    tax_mode = row["tax_mode"] or str(original_payload.get("taxMode") or "DEFAULT")
    cashier_name = cashier_override or row["cashier_name"] or "Не указан"
    timeout_minutes = max(5, get_int(settings, "TIMEOUT_MINUTES", 5))
    device_id = int(get_required(settings, "DEVICE_ID"))
    shop_id = int(get_required(settings, "SHOP_ID"))
    external_id = str(uuid.uuid4())

    payload = deepcopy(original_payload)
    payload["externalId"] = external_id
    payload["deviceId"] = device_id
    payload["shopId"] = shop_id
    payload["type"] = "REFUND"
    payload["timeout"] = timeout_minutes
    payload["taxMode"] = tax_mode
    payload["positions"] = selected_positions
    payload["payments"] = [{"sum": total_kopecks, "type": payment_type}]
    payload["attributes"] = attributes
    payload["total"] = {"priceSum": total_kopecks}
    legal_entity_tags = build_legal_entity_ffd_tags(str(buyer_type), buyer_legal_name, buyer_inn, settings)
    if legal_entity_tags and not payload.get("tags"):
        payload["tags"] = legal_entity_tags

    cashier_payload = build_cashier_payload(cashier_name)
    if cashier_payload:
        payload["cashier"] = cashier_payload
    else:
        payload.pop("cashier", None)

    positions = [position_to_receipt_position(pos) for pos in selected_positions]

    return ReceiptDraft(
        external_id=external_id,
        payment_type=str(payment_type),
        buyer_email=str(buyer_email) if buyer_email else None,
        buyer_phone=str(buyer_phone) if buyer_phone else None,
        buyer_type=str(buyer_type or "INDIVIDUAL"),
        buyer_legal_name=str(buyer_legal_name) if buyer_legal_name else None,
        buyer_inn=str(buyer_inn) if buyer_inn else None,
        tax_mode=str(tax_mode),
        cashier_name=str(cashier_name),
        positions=positions,
        total_kopecks=total_kopecks,
        payload=payload,
        parent_external_id=str(row["external_id"]),
        parent_receipt_id=str(row["receipt_id"] or "") or None,
        refund_mode=refund_mode,
        refunded_positions=selected_indices,
    )


def build_refund_draft_interactive(conn: sqlite3.Connection, settings: dict[str, str], cashier_name: Optional[str] = None) -> Optional[ReceiptDraft]:
    row = choose_receipt_for_refund(conn)
    if row is None:
        return None

    payload = load_json_object(row["request_json"])
    positions = payload_positions(payload)
    if not positions:
        print("В выбранном чеке нет сохраненных позиций. Возврат невозможен.")
        return None

    while True:
        print_saved_receipt_details(row)
        variant = choose_refund_variant(len(positions))
        if variant == "cancel":
            return None

        if variant == "full":
            selected_indices = list(range(1, len(positions) + 1))
            print_saved_receipt_details(row, selected_indices)
            answer = input("Возвращаем весь чек? Введите ДА для подтверждения: ").strip().lower()
            if answer in {"да", "д", "yes", "y"}:
                return build_refund_draft_from_row(row, settings, selected_indices, "full", cashier_name)
            print("Возврат всего чека не подтвержден. Возвращаюсь к выбору варианта.")
            continue

        if variant == "partial":
            selected_indices = choose_partial_positions(row, len(positions))
            if selected_indices is None:
                continue
            return build_refund_draft_from_row(row, settings, selected_indices, "partial", cashier_name)



def build_tbank_token(payload: dict[str, Any], password: str) -> str:
    """
    Формирует Token T-Банка.

    В токен входят только параметры корневого объекта.
    Вложенные объекты и массивы не участвуют.
    """
    root: dict[str, Any] = {}
    for key, value in payload.items():
        if key == "Token":
            continue
        if isinstance(value, (dict, list)):
            continue
        if value is None:
            continue
        root[key] = value
    root["Password"] = password
    concatenated = "".join(str(root[key]) for key in sorted(root))
    return hashlib.sha256(concatenated.encode("utf-8")).hexdigest()


def get_tbank_password(settings_path: Path, settings: dict[str, str], *, ask_if_missing: bool = True) -> str:
    """
    Возвращает пароль терминала T-Банка.

    Хранится как TBANK_PASSWORD_ENC, шифруется тем же переносимым механизмом,
    что и SMTP_PASSWORD_ENC: ключом от DREAMKAS_TOKEN.
    """
    token = settings.get("DREAMKAS_TOKEN", "").strip()
    encrypted = settings.get("TBANK_PASSWORD_ENC", "").strip()
    if encrypted:
        return decrypt_smtp_password(encrypted, token)

    plain_legacy = settings.get("TBANK_PASSWORD", "").strip()
    if plain_legacy:
        settings["TBANK_PASSWORD_ENC"] = encrypt_smtp_password(plain_legacy, token)
        settings.pop("TBANK_PASSWORD", None)
        save_settings(settings_path, settings)
        print("TBANK_PASSWORD найден в открытом виде, зашифрован и перенесен в TBANK_PASSWORD_ENC.")
        return plain_legacy

    if not ask_if_missing:
        return ""

    print("\nПароль терминала T-Банка не найден.")
    print("Введите пароль терминала один раз — он будет сохранен в settings.txt в зашифрованном виде.")
    password = getpass.getpass("TBANK_PASSWORD: ").strip()
    if password:
        settings["TBANK_PASSWORD_ENC"] = encrypt_smtp_password(password, token)
        settings.pop("TBANK_PASSWORD", None)
        save_settings(settings_path, settings)
        print("Пароль T-Банка зашифрован и сохранен как TBANK_PASSWORD_ENC.")
        return password
    return ""



def get_tbank_b2b_api_token(settings_path: Path, settings: dict[str, str], ask_if_missing: bool = True) -> str:
    """
    Возвращает Bearer token T-API для B2B/I2I СБП.

    Это отдельный токен T-API, не TerminalKey и не пароль интернет-эквайринга.
    """
    token_key = settings.get("DREAMKAS_TOKEN", "").strip()
    encrypted = settings.get("TBANK_B2B_API_TOKEN_ENC", "").strip()
    if encrypted:
        return decrypt_smtp_password(encrypted, token_key)

    plain_legacy = settings.get("TBANK_B2B_API_TOKEN", "").strip()
    if plain_legacy:
        settings["TBANK_B2B_API_TOKEN_ENC"] = encrypt_smtp_password(plain_legacy, token_key)
        settings.pop("TBANK_B2B_API_TOKEN", None)
        save_settings(settings_path, settings)
        print("TBANK_B2B_API_TOKEN найден в открытом виде, зашифрован и перенесен в TBANK_B2B_API_TOKEN_ENC.")
        return plain_legacy

    if not ask_if_missing:
        return ""

    print("\nДля СБП B2B/I2I нужен T-API Bearer token.")
    print("Он отличается от TerminalKey/Password интернет-эквайринга.")
    token = getpass.getpass("TBANK_B2B_API_TOKEN: ").strip()
    if token:
        settings["TBANK_B2B_API_TOKEN_ENC"] = encrypt_smtp_password(token, token_key)
        settings.pop("TBANK_B2B_API_TOKEN", None)
        save_settings(settings_path, settings)
        print("T-API token зашифрован и сохранен как TBANK_B2B_API_TOKEN_ENC.")
        return token
    return ""


def choose_sbp_mode_for_legal_buyer(settings: dict[str, str]) -> str:
    """
    Для юрлица/ИП при оплате через СБП спрашиваем:
    - обычный QR интернет-эквайринга;
    - СБП B2B/I2I через T-API.
    """
    print("\nПокупатель — юрлицо / ИП.")
    print("Выберите формат оплаты по СБП:")
    print("  1. Обычная СБП-оплата QR через интернет-эквайринг T-Банка")
    print("  2. СБП B2B/I2I — ссылка/QR для оплаты юрлицом или ИП через T-API")
    default = "2" if bool_setting(settings, "TBANK_B2B_SBP_ENABLED", False) else "1"

    while True:
        answer = input(f"Выберите вариант [1/2, Enter = {default}]: ").strip() or default
        if answer == "1":
            return "EACQ_SBP"
        if answer == "2":
            return "B2B_I2I"
        print("Введите 1 для обычной СБП-оплаты или 2 для СБП B2B/I2I.")


def normalize_b2b_vat(raw: str) -> str:
    value = str(raw or "").strip().replace("%", "")
    if value in {"0", "5", "7", "10", "20", "22"}:
        return value
    return "0"


def b2b_amount_rubles(amount_kopecks: int) -> float:
    """T-API B2B QR принимает сумму как число в рублях."""
    return round(int(amount_kopecks) / 100.0, 2)


def extract_b2b_qr_id(response: dict[str, Any]) -> Optional[str]:
    for key in ("qrId", "id", "linkId", "paymentId", "paymentLinkId"):
        value = response.get(key)
        if value:
            return str(value)
    data = response.get("data")
    if isinstance(data, dict):
        for key in ("qrId", "id", "linkId", "paymentId", "paymentLinkId"):
            value = data.get(key)
            if value:
                return str(value)
    return None


def extract_b2b_payment_link(response: dict[str, Any]) -> Optional[str]:
    candidates = [
        "payload", "Payload", "qrPayload", "qrUrl", "url", "link", "paymentUrl",
        "paymentLink", "redirectUrl", "shortUrl", "deeplink", "Data", "data",
    ]
    for key in candidates:
        value = response.get(key)
        if isinstance(value, str) and value.strip():
            return value.strip()
    data = response.get("data")
    if isinstance(data, dict):
        for key in candidates:
            value = data.get(key)
            if isinstance(value, str) and value.strip():
                return value.strip()
    return None


def extract_b2b_status(response: dict[str, Any]) -> str:
    for key in ("status", "Status", "state", "State", "paymentStatus", "PaymentStatus"):
        value = response.get(key)
        if value:
            return str(value).upper()
    data = response.get("data")
    if isinstance(data, dict):
        for key in ("status", "Status", "state", "State", "paymentStatus", "PaymentStatus"):
            value = data.get(key)
            if value:
                return str(value).upper()
    return "UNKNOWN"



def tbank_request(endpoint: str, payload: dict[str, Any], settings: dict[str, str], settings_path: Path) -> dict[str, Any]:
    terminal_key = settings.get("TBANK_TERMINAL_KEY", "").strip()
    if not terminal_key:
        raise DreamkasError("Для оплаты по СБП нужно настроить TBANK_TERMINAL_KEY.")

    password = get_tbank_password(settings_path, settings, ask_if_missing=True)
    if not password:
        raise DreamkasError("Для оплаты по СБП нужен пароль терминала T-Банка.")

    body = dict(payload)
    body["TerminalKey"] = terminal_key
    body["Token"] = build_tbank_token(body, password)

    url = f"https://securepay.tinkoff.ru/v2/{endpoint}"
    timeout_seconds = max(5, get_int(settings, "TBANK_REQUEST_TIMEOUT_SECONDS", 30))
    log_json(f"REQUEST POST {url}", {k: ("***" if k == "Token" else v) for k, v in body.items()})

    try:
        response = requests.post(url, json=body, timeout=timeout_seconds)
    except requests.RequestException as exc:
        logging.exception("T-Bank network error for %s", endpoint)
        raise DreamkasError(f"Сетевая ошибка T-Банка {endpoint}: {exc}") from exc

    try:
        data = response.json()
    except Exception:
        data = {"raw": response.text}

    log_json(f"RESPONSE {response.status_code} POST {url}", data)
    if response.status_code < 200 or response.status_code >= 300:
        raise DreamkasError(f"T-Банк API вернул HTTP {response.status_code}: {data}")
    if isinstance(data, dict) and data.get("Success") is False:
        msg = data.get("Message") or data.get("Details") or data.get("ErrorCode") or data
        raise DreamkasError(f"T-Банк API вернул ошибку: {msg}")
    if not isinstance(data, dict):
        raise DreamkasError("T-Банк API вернул неожиданный ответ")
    return data


def create_tbank_sbp_payment(draft: ReceiptDraft, settings: dict[str, str], settings_path: Path) -> dict[str, Any]:
    description_prefix = settings.get("TBANK_DESCRIPTION_PREFIX", "Оплата заказа").strip() or "Оплата заказа"
    description = f"{description_prefix} {draft.external_id}"[:140]
    due_minutes = max(1, get_int(settings, "TBANK_PAYMENT_TIMEOUT_MINUTES", 10))
    due_date = (datetime.now() + timedelta(minutes=due_minutes)).astimezone().replace(microsecond=0).isoformat()
    payload = {
        "Amount": draft.total_kopecks,
        "OrderId": draft.external_id,
        "Description": description,
        "PayType": settings.get("TBANK_PAY_TYPE", "O").strip().upper() or "O",
        "RedirectDueDate": due_date,
    }
    return tbank_request("Init", payload, settings, settings_path)


def get_tbank_qr(payment_id: str, settings: dict[str, str], settings_path: Path) -> dict[str, Any]:
    payload = {
        "PaymentId": str(payment_id),
        "DataType": settings.get("TBANK_QR_DATA_TYPE", "PAYLOAD").strip().upper() or "PAYLOAD",
    }
    return tbank_request("GetQr", payload, settings, settings_path)


def get_tbank_state(payment_id: str, settings: dict[str, str], settings_path: Path) -> dict[str, Any]:
    return tbank_request("GetState", {"PaymentId": str(payment_id)}, settings, settings_path)



def tbank_b2b_request(
    method: str,
    path: str,
    *,
    settings: dict[str, str],
    settings_path: Path,
    json_body: Optional[dict[str, Any]] = None,
    query: str = "",
) -> dict[str, Any]:
    """
    Запрос к T-API для СБП B2B/I2I.
    """
    token = get_tbank_b2b_api_token(settings_path, settings, ask_if_missing=True)
    if not token:
        raise DreamkasError("Для СБП B2B/I2I нужен TBANK_B2B_API_TOKEN.")

    base_url = "https://business.tbank.ru/openapi"
    url = base_url + path + query
    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json",
        "Accept": "application/json",
        "X-Request-Id": str(uuid.uuid4()),
    }
    timeout_seconds = max(5, get_int(settings, "TBANK_REQUEST_TIMEOUT_SECONDS", 30))

    log_json(f"REQUEST {method} {url}", json_body or {})
    try:
        response = requests.request(method, url, headers=headers, json=json_body, timeout=timeout_seconds)
    except requests.RequestException as exc:
        logging.exception("T-Bank B2B network error for %s %s", method, url)
        raise DreamkasError(f"Сетевая ошибка T-Банк B2B/I2I: {exc}") from exc

    body = parse_api_response(response)
    log_json(f"RESPONSE {response.status_code} {method} {url}", body)
    if response.status_code < 200 or response.status_code >= 300:
        raise DreamkasError(f"T-Банк B2B/I2I API вернул HTTP {response.status_code}: {body}")

    if isinstance(body, dict):
        return body
    return {"raw": body}


def create_tbank_b2b_sbp_link(draft: ReceiptDraft, settings: dict[str, str], settings_path: Path) -> dict[str, Any]:
    account_number = re.sub(r"\D", "", settings.get("TBANK_B2B_ACCOUNT_NUMBER", ""))
    if not re.fullmatch(r"(\d{20}|\d{22})", account_number or ""):
        raise DreamkasError("Для СБП B2B/I2I нужно указать TBANK_B2B_ACCOUNT_NUMBER — расчетный счет 20 или 22 цифры.")

    ttl_days = max(1, get_int(settings, "TBANK_B2B_TTL_DAYS", 1))
    purpose_prefix = settings.get("TBANK_B2B_PURPOSE_PREFIX", "Оплата по счету").strip() or "Оплата по счету"
    purpose_parts = [purpose_prefix, draft.external_id]
    if draft.buyer_legal_name:
        purpose_parts.append(str(draft.buyer_legal_name))
    purpose = " ".join(purpose_parts)[:210]

    payload: dict[str, Any] = {
        "accountNumber": account_number,
        "sum": b2b_amount_rubles(draft.total_kopecks),
        "purpose": purpose,
        "ttl": ttl_days,
        "vat": normalize_b2b_vat(settings.get("TBANK_B2B_VAT", "0")),
    }

    redirect_url = settings.get("TBANK_B2B_REDIRECT_URL", "").strip()
    if redirect_url:
        payload["redirectUrl"] = redirect_url

    return tbank_b2b_request(
        "POST",
        "/api/v1/b2b/qr/onetime",
        settings=settings,
        settings_path=settings_path,
        json_body=payload,
    )


def get_tbank_b2b_sbp_info(qr_id: str, settings: dict[str, str], settings_path: Path, with_image: bool = False) -> dict[str, Any]:
    query = "?withImage=true" if with_image else ""
    return tbank_b2b_request(
        "GET",
        f"/api/v1/b2b/qr/{qr_id}/info",
        settings=settings,
        settings_path=settings_path,
        query=query,
    )


def run_tbank_b2b_sbp_payment(
    *,
    conn: sqlite3.Connection,
    draft: ReceiptDraft,
    settings: dict[str, str],
    settings_path: Path,
    paths: dict[str, Path],
) -> SbpPaymentResult:
    """
    Создает B2B/I2I СБП-ссылку/QR через T-API и ждёт оплату.
    """
    if not bool_setting(settings, "TBANK_B2B_SBP_ENABLED", False):
        raise DreamkasError("СБП B2B/I2I выбрано, но TBANK_B2B_SBP_ENABLED = 0. Включите его в настройках.")

    print("\nСоздаю СБП B2B/I2I ссылку в T-Банке...")
    create_response = create_tbank_b2b_sbp_link(draft, settings, settings_path)
    qr_id = extract_b2b_qr_id(create_response) or draft.external_id
    qr_payload = extract_b2b_payment_link(create_response)

    info_response: dict[str, Any] = create_response
    if not qr_payload and qr_id:
        try:
            info_response = get_tbank_b2b_sbp_info(str(qr_id), settings, settings_path, with_image=True)
            qr_payload = extract_b2b_payment_link(info_response)
        except Exception:
            logging.exception("Could not fetch B2B QR info immediately")

    if not qr_payload:
        raise DreamkasError(f"T-Банк B2B/I2I не вернул ссылку/QR payload: {create_response}")

    sbp_qr_name = f"SBP-B2B({datetime.now().strftime('%Y-%m-%d-%H-%M-%S')}).png"
    sbp_qr_path = paths["receipts_qr"] / sbp_qr_name
    save_qr(qr_payload, sbp_qr_path)
    print(f"QR СБП B2B/I2I сохранен: {sbp_qr_path}")
    print_terminal_qr(qr_payload)

    db_update(
        conn,
        draft.external_id,
        sbp_payment_id=str(qr_id),
        sbp_order_id=draft.external_id,
        sbp_status=extract_b2b_status(create_response),
        sbp_qr_payload=qr_payload,
        sbp_qr_path=str(sbp_qr_path),
        sbp_json=json.dumps({"b2b_create": create_response, "b2b_info": info_response}, ensure_ascii=False, indent=2),
    )

    print("\nОжидаю оплату СБП B2B/I2I...")
    print("После успешной оплаты чек будет отправлен в Dreamkas автоматически.")
    print_cancel_hint()

    poll_interval = max(3, get_int(settings, "TBANK_POLL_INTERVAL_SECONDS", 5))
    timeout_minutes = max(1, get_int(settings, "TBANK_PAYMENT_TIMEOUT_MINUTES", 10))
    deadline = datetime.now() + timedelta(minutes=timeout_minutes)

    success_statuses = {"PAID", "CONFIRMED", "SUCCESS", "SUCCEEDED", "PAYED", "COMPLETED", "EXECUTED"}
    failure_statuses = {"CANCELED", "CANCELLED", "REJECTED", "EXPIRED", "DEADLINE_EXPIRED", "FAILED", "ERROR"}

    status = extract_b2b_status(info_response)
    last_info = info_response

    while datetime.now() < deadline:
        raise_if_operator_cancelled("ожидание оплаты СБП B2B/I2I")
        if qr_id:
            last_info = get_tbank_b2b_sbp_info(str(qr_id), settings, settings_path, with_image=False)
            status = extract_b2b_status(last_info)
        db_update(
            conn,
            draft.external_id,
            sbp_status=f"B2B_{status}",
            sbp_json=json.dumps({"b2b_create": create_response, "b2b_info": last_info}, ensure_ascii=False, indent=2),
        )
        if status in success_statuses:
            finish_inline_status()
            print(f"Оплата СБП B2B/I2I подтверждена. Статус: {status}")
            return SbpPaymentResult(str(qr_id), draft.external_id, f"B2B_{status}", qr_payload, str(sbp_qr_path), create_response, {"b2b_qr": qr_payload}, last_info)
        if status in failure_statuses:
            finish_inline_status()
            raise DreamkasError(f"Платеж СБП B2B/I2I завершился неуспешно. Статус: {status}")

        print_inline_status(f"Статус оплаты B2B/I2I: {status or 'UNKNOWN'}. Следующая проверка через {poll_interval} сек.")
        interruptible_sleep(poll_interval, "ожидание оплаты СБП B2B/I2I")

    finish_inline_status()
    raise DreamkasError(f"Истекло время ожидания оплаты СБП B2B/I2I ({timeout_minutes} мин). Чек в Dreamkas не отправлен.")



def cancel_tbank_payment(payment_id: str, amount_kopecks: Optional[int], settings: dict[str, str], settings_path: Path) -> dict[str, Any]:
    """
    Отмена / возврат платежа в T-Банке через /v2/Cancel.

    Для полного и частичного возврата передаем Amount в копейках.
    Чеки формируются отдельно через Dreamkas, поэтому Receipt в T-Банк не передаем.
    """
    payload: dict[str, Any] = {"PaymentId": str(payment_id)}
    if amount_kopecks is not None and int(amount_kopecks) > 0:
        payload["Amount"] = int(amount_kopecks)
    return tbank_request("Cancel", payload, settings, settings_path)


def get_receipt_row_by_external_id(conn: sqlite3.Connection, external_id: Optional[str]) -> Optional[sqlite3.Row]:
    if not external_id:
        return None
    return conn.execute(
        "SELECT * FROM receipts WHERE external_id = ? LIMIT 1",
        (str(external_id),),
    ).fetchone()


def maybe_offer_tbank_refund_after_fiscal_refund(
    *,
    conn: sqlite3.Connection,
    draft: ReceiptDraft,
    settings: dict[str, str],
    settings_path: Path,
) -> None:
    """
    После успешной фискализации чека возврата предлагает вернуть деньги через T-Банк,
    если исходная продажа была оплачена через СБП и в SQLite сохранен sbp_payment_id.
    """
    if str(draft.payload.get("type") or "").upper() != "REFUND":
        return

    parent = get_receipt_row_by_external_id(conn, draft.parent_external_id)
    if parent is None:
        return

    parent_payment_id = str(parent["sbp_payment_id"] or "").strip()
    if not parent_payment_id:
        return

    parent_sbp_status = str(parent["sbp_status"] or "").strip()
    print("\nИсходный чек был оплачен через Т-Банк СБП.")
    print(f"PaymentId исходной оплаты: {parent_payment_id}")
    if parent_sbp_status:
        print(f"Статус исходной оплаты: {parent_sbp_status}")
    print(f"Сумма возврата денег: {format_money(draft.total_kopecks)}")

    answer = input("Отправить возврат денег через Т-Банк СБП? Введите ДА: ").strip().lower()
    if answer not in {"да", "д", "yes", "y"}:
        db_update(
            conn,
            draft.external_id,
            sbp_status="TBANK_REFUND_SKIPPED",
            sbp_json=json.dumps(
                {
                    "money_refund": "skipped_by_operator",
                    "parent_sbp_payment_id": parent_payment_id,
                    "amount_kopecks": draft.total_kopecks,
                    "skipped_at": datetime.now().isoformat(timespec="seconds"),
                },
                ensure_ascii=False,
                indent=2,
            ),
        )
        print("Возврат денег через Т-Банк пропущен оператором.")
        return

    print("Отправляю запрос возврата денег в Т-Банк...")
    try:
        cancel_response = cancel_tbank_payment(parent_payment_id, draft.total_kopecks, settings, settings_path)
        refund_status = str(cancel_response.get("Status") or cancel_response.get("Success") or "UNKNOWN")
        db_update(
            conn,
            draft.external_id,
            sbp_payment_id=parent_payment_id,
            sbp_status=f"TBANK_REFUND_{refund_status}",
            sbp_json=json.dumps(
                {
                    "money_refund": "sent",
                    "parent_sbp_payment_id": parent_payment_id,
                    "amount_kopecks": draft.total_kopecks,
                    "cancel_response": cancel_response,
                    "sent_at": datetime.now().isoformat(timespec="seconds"),
                },
                ensure_ascii=False,
                indent=2,
            ),
        )
        print(f"Запрос возврата денег отправлен в Т-Банк. Статус: {refund_status}")
    except Exception as exc:
        db_update(
            conn,
            draft.external_id,
            sbp_payment_id=parent_payment_id,
            sbp_status="TBANK_REFUND_ERROR",
            sbp_json=json.dumps(
                {
                    "money_refund": "error",
                    "parent_sbp_payment_id": parent_payment_id,
                    "amount_kopecks": draft.total_kopecks,
                    "error": str(exc),
                    "failed_at": datetime.now().isoformat(timespec="seconds"),
                },
                ensure_ascii=False,
                indent=2,
            ),
        )
        print("\nВнимание: фискальный чек возврата уже создан, но возврат денег через Т-Банк не прошел.")
        print(f"Ошибка Т-Банка: {exc}")
        logging.exception("T-Bank money refund failed")



def extract_tbank_qr_payload(qr_response: dict[str, Any]) -> Optional[str]:
    for key in ("Data", "Payload", "QrCode", "QR", "Url"):
        value = qr_response.get(key)
        if value:
            return str(value)
    return None


def print_terminal_qr(data: str) -> None:
    qr = qrcode.QRCode(border=1)
    qr.add_data(data)
    qr.make(fit=True)
    matrix = qr.get_matrix()
    print("\nQR-код СБП для оплаты:")
    print("(покупатель должен отсканировать его банковским приложением)")
    for row in matrix:
        print("".join("██" if cell else "  " for cell in row))


def run_tbank_sbp_payment(
    *,
    conn: sqlite3.Connection,
    draft: ReceiptDraft,
    settings: dict[str, str],
    settings_path: Path,
    paths: dict[str, Path],
) -> SbpPaymentResult:
    """
    Создаёт платеж СБП T-Банка, показывает QR в терминале,
    сохраняет QR в PNG и ждёт успешной оплаты.
    """
    if draft.total_kopecks < 1000:
        raise DreamkasError("Минимальная сумма операции по СБП в T-Банке — 10 рублей.")
    if not bool_setting(settings, "TBANK_SBP_ENABLED", False):
        raise DreamkasError("Оплата СБП выбрана, но TBANK_SBP_ENABLED = 0. Включите СБП в настройках.")

    print("\nСоздаю платеж СБП в T-Банке...")
    init_response = create_tbank_sbp_payment(draft, settings, settings_path)
    payment_id = str(init_response.get("PaymentId") or "")
    order_id = str(init_response.get("OrderId") or draft.external_id)
    if not payment_id:
        raise DreamkasError(f"T-Банк Init не вернул PaymentId: {init_response}")

    db_update(conn, draft.external_id, sbp_payment_id=payment_id, sbp_order_id=order_id,
              sbp_status=str(init_response.get("Status") or "INIT"),
              sbp_json=json.dumps({"init": init_response}, ensure_ascii=False, indent=2))
    print(f"Платеж создан. PaymentId: {payment_id}")

    qr_response = get_tbank_qr(payment_id, settings, settings_path)
    qr_payload = extract_tbank_qr_payload(qr_response)
    if not qr_payload:
        raise DreamkasError(f"T-Банк GetQr не вернул Payload/Data для QR: {qr_response}")

    sbp_qr_name = f"SBP({datetime.now().strftime('%Y-%m-%d-%H-%M-%S')}).png"
    sbp_qr_path = paths["receipts_qr"] / sbp_qr_name
    save_qr(qr_payload, sbp_qr_path)
    print(f"QR СБП сохранен: {sbp_qr_path}")
    print_terminal_qr(qr_payload)
    db_update(conn, draft.external_id, sbp_qr_payload=qr_payload, sbp_qr_path=str(sbp_qr_path),
              sbp_json=json.dumps({"init": init_response, "qr": qr_response}, ensure_ascii=False, indent=2))

    print("\nОжидаю оплату СБП...")
    print("После успешной оплаты чек будет отправлен в Dreamkas автоматически.")
    print_cancel_hint()
    poll_interval = max(3, get_int(settings, "TBANK_POLL_INTERVAL_SECONDS", 5))
    timeout_minutes = max(1, get_int(settings, "TBANK_PAYMENT_TIMEOUT_MINUTES", 10))
    deadline = datetime.now() + timedelta(minutes=timeout_minutes)

    success_statuses = {"CONFIRMED", "AUTHORIZED"}
    failure_statuses = {"REJECTED", "CANCELED", "CANCELLED", "DEADLINE_EXPIRED", "AUTH_FAIL"}
    state_response: dict[str, Any] = init_response
    status = str(init_response.get("Status") or "NEW").upper()

    while datetime.now() < deadline:
        raise_if_operator_cancelled("ожидание оплаты СБП")
        state_response = get_tbank_state(payment_id, settings, settings_path)
        status = str(state_response.get("Status") or "").upper()
        db_update(conn, draft.external_id, sbp_status=status,
                  sbp_json=json.dumps({"init": init_response, "qr": qr_response, "state": state_response}, ensure_ascii=False, indent=2))
        if status in success_statuses:
            finish_inline_status()
            print(f"Оплата СБП подтверждена. Статус: {status}")
            return SbpPaymentResult(payment_id, order_id, status, qr_payload, str(sbp_qr_path), init_response, qr_response, state_response)
        if status in failure_statuses:
            finish_inline_status()
            raise DreamkasError(f"Платеж СБП завершился неуспешно. Статус: {status}")
        print_inline_status(f"Статус оплаты: {status or 'UNKNOWN'}. Следующая проверка через {poll_interval} сек.")
        interruptible_sleep(poll_interval, "ожидание оплаты СБП")

    finish_inline_status()
    raise DreamkasError(f"Истекло время ожидания оплаты СБП ({timeout_minutes} мин). Чек в Dreamkas не отправлен.")


def confirm_external_terminal_payment(conn: sqlite3.Connection, draft: ReceiptDraft) -> bool:
    """
    Для безнала через внешний терминал программа не может проверить оплату сама.

    Поэтому оператор должен вручную подтвердить, что платеж на внешнем терминале
    успешно принят. Только после этого чек отправляется в Dreamkas.
    """
    if str(draft.payload.get("type") or "").upper() != "SALE":
        return True
    if draft.payment_type != "CASHLESS":
        return True

    print("\nПОДТВЕРЖДЕНИЕ ОПЛАТЫ ВНЕШНИМ ТЕРМИНАЛОМ")
    print("-" * 72)
    print(f"Сумма к оплате: {format_money(draft.total_kopecks)}")
    print("Проведите оплату на внешнем банковском терминале.")
    print("Чек в Dreamkas будет отправлен только после подтверждения оплаты оператором.")
    print("-" * 72)

    answer = input("Оплата на внешнем терминале успешно принята? Введите ДА: ").strip().lower()
    if answer in {"да", "д", "yes", "y"}:
        db_update(
            conn,
            draft.external_id,
            sbp_status="EXTERNAL_TERMINAL_CONFIRMED",
            sbp_json=json.dumps(
                {
                    "provider": "EXTERNAL_TERMINAL",
                    "confirmed_at": datetime.now().isoformat(timespec="seconds"),
                    "amount_kopecks": draft.total_kopecks,
                    "operator_confirmation": True,
                },
                ensure_ascii=False,
                indent=2,
            ),
        )
        return True

    db_update(
        conn,
        draft.external_id,
        status="PAYMENT_NOT_CONFIRMED",
        completed_at=datetime.now().isoformat(timespec="seconds"),
        error_code="PAYMENT_NOT_CONFIRMED",
        error_message="Оператор не подтвердил оплату на внешнем терминале",
        sbp_status="EXTERNAL_TERMINAL_NOT_CONFIRMED",
    )
    return False


def maybe_run_payment_before_fiscalization(
    *,
    conn: sqlite3.Connection,
    draft: ReceiptDraft,
    settings: dict[str, str],
    settings_path: Path,
    paths: dict[str, Path],
) -> Optional[SbpPaymentResult]:
    if str(draft.payload.get("type") or "").upper() != "SALE":
        return None

    if draft.payment_type == "SBP":
        if draft.buyer_type == "LEGAL_ENTITY":
            sbp_mode = choose_sbp_mode_for_legal_buyer(settings)
            if sbp_mode == "B2B_I2I":
                return run_tbank_b2b_sbp_payment(conn=conn, draft=draft, settings=settings, settings_path=settings_path, paths=paths)
        return run_tbank_sbp_payment(conn=conn, draft=draft, settings=settings, settings_path=settings_path, paths=paths)

    if draft.payment_type == "CASHLESS":
        if not confirm_external_terminal_payment(conn, draft):
            raise DreamkasError("Оплата на внешнем терминале не подтверждена. Чек в Dreamkas не отправлен.")

    return None


def submit_draft(
    *,
    conn: sqlite3.Connection,
    session: requests.Session,
    api_base_url: str,
    draft: ReceiptDraft,
    poll_interval: int,
    request_timeout: int,
    workdir: Path,
    action_label: str,
    settings: dict[str, str],
    settings_path: Path,
    excel_path: Optional[Path] = None,
) -> int:
    print_precheck(draft)

    errors, warnings = validate_draft_for_send(draft)
    if not print_validation_result(errors, warnings):
        print("\nОтправка отменена из-за ошибок или неподтвержденных предупреждений.")
        wait_key()
        return 1

    if not confirm_no_recent_duplicate(conn, draft, settings):
        print("\nОтправка отменена: возможный дубль.")
        wait_key()
        return 0

    answer = input(f"\n{action_label}? Введите ДА для отправки: ").strip().lower()
    if answer not in {"да", "д", "yes", "y"}:
        print("\nОтменено пользователем. Чек НЕ отправлен.")
        wait_key()
        return 0

    paths = ensure_work_folders(workdir)
    db_insert_draft(conn, draft, excel_path)
    print(f"\nЧерновик и полный предчек сохранены в SQLite: {paths['db'] / DB_FILE}")
    logging.info(
        "Draft saved: external_id=%s type=%s amount=%s",
        draft.external_id,
        draft.payload.get("type"),
        draft.total_kopecks,
    )

    try:
        sbp_result = maybe_run_payment_before_fiscalization(
            conn=conn,
            draft=draft,
            settings=settings,
            settings_path=settings_path,
            paths=paths,
        )
        if sbp_result:
            print("Оплата СБП успешна. Продолжаю фискализацию чека.")
    except OperatorCancelled as cancel_exc:
        finish_inline_status()
        db_update(
            conn,
            draft.external_id,
            status="CANCELLED_BY_OPERATOR",
            completed_at=datetime.now().isoformat(timespec="seconds"),
            error_code="CANCELLED_BY_OPERATOR",
            error_message=str(cancel_exc),
        )
        print("\nОжидание оплаты отменено оператором.")
        print("Чек в Dreamkas НЕ отправлен.")
        print("Возврат в главное меню.")
        wait_key("Нажмите любую клавишу для возврата в главное меню...")
        return 0
    except Exception as sbp_exc:
        db_update(
            conn,
            draft.external_id,
            status="SBP_ERROR",
            completed_at=datetime.now().isoformat(timespec="seconds"),
            error_code="SBP_ERROR",
            error_message=str(sbp_exc),
        )
        print("\nОШИБКА ОПЛАТЫ")
        print(str(sbp_exc))
        print("Чек в Dreamkas НЕ отправлен.")
        logging.exception("Payment step failed")
        wait_key()
        return 1

    print("\nОтправляю чек в Dreamkas...")
    operation = send_receipt(session, api_base_url, draft, request_timeout)
    operation_id = str(operation.get("id") or draft.external_id)
    status = str(operation.get("status") or "PENDING")
    db_update(
        conn,
        draft.external_id,
        operation_id=operation.get("id"),
        status=status,
        sent_at=datetime.now().isoformat(timespec="seconds"),
        operation_json=json.dumps(operation, ensure_ascii=False, indent=2),
    )
    print(f"Операция создана: {operation_id}, статус: {status}")

    final_operation = operation
    print_cancel_hint()
    try:
        while True:
            status = str(final_operation.get("status") or "").upper()
            if status in {"SUCCESS", "ERROR"}:
                break
            print(f"Статус: {status or 'UNKNOWN'}. Следующая проверка через {poll_interval} сек.")
            interruptible_sleep(poll_interval, "ожидание ответа Dreamkas")
            raise_if_operator_cancelled("ожидание ответа Dreamkas")
            final_operation = get_operation(session, api_base_url, operation_id, request_timeout)
            db_update(
                conn,
                draft.external_id,
                status=str(final_operation.get("status") or "UNKNOWN"),
                operation_json=json.dumps(final_operation, ensure_ascii=False, indent=2),
            )
    except OperatorCancelled as cancel_exc:
        db_update(
            conn,
            draft.external_id,
            status="CANCELLED_BY_OPERATOR",
            completed_at=datetime.now().isoformat(timespec="seconds"),
            error_code="CANCELLED_BY_OPERATOR",
            error_message=str(cancel_exc),
            operation_json=json.dumps(final_operation, ensure_ascii=False, indent=2),
        )
        print("\nОжидание ответа Dreamkas отменено оператором.")
        print("Возврат в главное меню.")
        print("Важно: задание уже было отправлено в Dreamkas и может продолжить выполняться на стороне кассы.")
        print("Перед повторной пробивкой проверьте статус этого чека, чтобы не создать дубль.")
        wait_key("Нажмите любую клавишу для возврата в главное меню...")
        return 0

    status = str(final_operation.get("status") or "").upper()
    if status == "ERROR":
        err = (final_operation.get("data") or {}).get("error") or {}
        error_code = str(err.get("code") or "")
        error_message = str(err.get("message") or "Ошибка без описания")
        db_update(
            conn,
            draft.external_id,
            status="ERROR",
            completed_at=datetime.now().isoformat(timespec="seconds"),
            operation_json=json.dumps(final_operation, ensure_ascii=False, indent=2),
            error_code=error_code,
            error_message=error_message,
        )
        print("\nОШИБКА ФИСКАЛИЗАЦИИ")
        print(f"Код: {error_code or '-'}")
        print(f"Сообщение: {error_message}")
        logging.error("Fiscalization error %s: %s", error_code, error_message)
        wait_key()
        return 1

    receipt_id = (final_operation.get("data") or {}).get("receiptId")
    if not receipt_id:
        raise DreamkasError("Операция SUCCESS, но в data.receiptId нет идентификатора чека")

    print(f"\nЧек успешно фискализирован. Receipt ID: {receipt_id}")
    receipt = get_receipt(session, api_base_url, str(receipt_id), request_timeout)

    now_name = datetime.now().strftime("%Y-%m-%d-%H-%M-%S")
    prefix = "Refund" if str(draft.payload.get("type") or "").upper() == "REFUND" else "Receipt"
    base_name = f"{prefix}({now_name})"
    txt_path = paths["receipts_txt"] / f"{base_name}.txt"
    qr_path = paths["receipts_qr"] / f"{base_name}.png"
    pdf_path = paths["receipts_pdf"] / f"{base_name}.pdf"

    qr_data: Optional[str] = None
    try:
        qr_data = build_fns_qr_data(receipt)
        save_qr(qr_data, qr_path)
        qr_path_to_db = str(qr_path)
        print(f"QR-код сохранен: {qr_path}")
    except Exception as qr_exc:
        qr_path_to_db = None
        print(f"Внимание: QR-код ФНС не удалось сформировать: {qr_exc}")
        logging.exception("Could not save QR code")

    save_receipt_txt(txt_path, draft, final_operation, receipt, qr_data)
    print(f"TXT-чек сохранен: {txt_path}")

    pdf_path_to_db = None
    if bool_setting(settings, "PDF_ENABLED", True):
        try:
            save_receipt_pdf(pdf_path, draft, final_operation, receipt, qr_path if qr_path.exists() else None, qr_data)
            pdf_path_to_db = str(pdf_path)
            print(f"PDF-чек сохранен: {pdf_path}")
        except Exception as pdf_exc:
            print(f"Внимание: PDF-чек не удалось сформировать: {pdf_exc}")
            logging.exception("Could not save PDF receipt")

    send_receipt_email(
        settings_path,
        settings,
        draft,
        receipt,
        txt_path if txt_path.exists() else None,
        Path(pdf_path_to_db) if pdf_path_to_db else None,
        qr_path if qr_path.exists() else None,
    )

    db_update(
        conn,
        draft.external_id,
        status="SUCCESS",
        completed_at=datetime.now().isoformat(timespec="seconds"),
        receipt_id=str(receipt_id),
        operation_json=json.dumps(final_operation, ensure_ascii=False, indent=2),
        receipt_json=json.dumps(receipt, ensure_ascii=False, indent=2),
        receipt_txt_path=str(txt_path),
        qr_path=qr_path_to_db,
        pdf_path=pdf_path_to_db,
    )

    maybe_offer_tbank_refund_after_fiscal_refund(
        conn=conn,
        draft=draft,
        settings=settings,
        settings_path=settings_path,
    )

    if excel_path is not None and bool_setting(settings, "CLEAR_EXCEL_AFTER_SUCCESS", True):
        try:
            clear_excel_template_after_success(excel_path)
            print(f"Excel-шаблон очищен после успешной фискализации: {excel_path}")
        except PermissionError:
            print("Внимание: Excel-шаблон не удалось очистить — файл открыт или заблокирован.")
            print("Закройте Excel и очистите файл вручную либо повторите очистку позже.")
            logging.exception("Could not clear Excel template because it is locked")
        except Exception as clear_exc:
            print(f"Внимание: Excel-шаблон не удалось очистить: {clear_exc}")
            logging.exception("Could not clear Excel template")

    logging.info("Receipt completed: external_id=%s receipt_id=%s", draft.external_id, receipt_id)
    print("\nГотово.")
    wait_key("Нажмите любую клавишу для возврата в главное меню...")
    return 0


def main() -> int:
    enable_ansi_colors()
    parser = argparse.ArgumentParser(description="Отправка и возврат чеков Dreamkas")
    parser.add_argument("--settings", default="settings.txt", help="TXT-файл настроек TOKEN_KEY = VALUE")
    parser.add_argument("--excel", default="dreamkas_receipt_template.xlsx", help="Excel-файл с новым чеком")
    args = parser.parse_args()

    workdir = get_app_dir()
    paths = ensure_work_folders(workdir)
    configure_logging(paths["logs"])
    settings_path = resolve_app_file(Path(args.settings))
    excel_path = ensure_excel_template_available(Path(args.excel))
    db_path = paths["db"] / DB_FILE

    try:
        settings = load_settings(settings_path)
        settings = ensure_token(settings_path, settings)
        token = get_required(settings, "DREAMKAS_TOKEN")
        api_base_url = settings.get("API_BASE_URL", API_DEFAULT_BASE_URL).strip() or API_DEFAULT_BASE_URL
        request_timeout = max(5, get_int(settings, "REQUEST_TIMEOUT_SECONDS", 30))

        session = requests.Session()
        session.headers.update(dreamkas_headers(token))

        try:
            settings = select_shop_and_device(settings_path, settings, session, api_base_url, request_timeout)
        except DreamkasError as exc:
            if is_token_auth_error(exc):
                clear_smtp_settings(settings_path, settings, str(exc))
                wait_key("Нажмите любую клавишу для выхода...")
                return 1
            raise

        # Мягкая миграция старого открытого SMTP_PASSWORD, если он остался в settings.txt.
        if settings.get("SMTP_PASSWORD"):
            get_smtp_password(settings_path, settings, ask_if_missing=False)

        conn = init_db(db_path)

        session_cashier_name = prompt_session_cashier(settings_path, settings)

        while True:
            poll_interval = max(10, get_int(settings, "POLL_INTERVAL_SECONDS", 20))
            mode = choose_work_mode()
            if mode == "exit":
                print("\nВыход без действий.")
                return 0

            if mode == "history":
                show_history_menu(conn)
                continue

            if mode == "settings":
                settings = settings_menu(settings_path, settings, session, api_base_url, request_timeout)
                token = get_required(settings, "DREAMKAS_TOKEN")
                session.headers.update(dreamkas_headers(token))
                continue

            if mode == "sale":
                try:
                    sale_meta = prompt_sale_metadata(settings_path, settings, session_cashier_name)
                    open_excel_for_user_fill(excel_path)
                    draft = read_excel_receipt(excel_path, settings, sale_meta)
                except OperatorCancelled as cancel_exc:
                    print(f"\nОперация отменена. {cancel_exc}")
                    wait_key("Нажмите любую клавишу для возврата в главное меню...")
                    continue

                submit_draft(
                    conn=conn,
                    session=session,
                    api_base_url=api_base_url,
                    draft=draft,
                    poll_interval=poll_interval,
                    request_timeout=request_timeout,
                    workdir=workdir,
                    action_label="Отправить чек на фискализацию",
                    settings=settings,
                    settings_path=settings_path,
                    excel_path=excel_path,
                )
                continue

            if mode == "refund":
                draft = build_refund_draft_interactive(conn, settings, session_cashier_name)
                if draft is None:
                    print("\nВозврат отменен. Чек НЕ отправлен.")
                    wait_key("Нажмите любую клавишу для возврата в главное меню...")
                    continue
                submit_draft(
                    conn=conn,
                    session=session,
                    api_base_url=api_base_url,
                    draft=draft,
                    poll_interval=poll_interval,
                    request_timeout=request_timeout,
                    workdir=workdir,
                    action_label="Отправить чек возврата",
                    settings=settings,
                    settings_path=settings_path,
                )
                continue

            raise ValueError("Неизвестный режим работы")

    except Exception as exc:
        logging.error("Fatal error: %s\n%s", exc, traceback.format_exc())
        print("\nОШИБКА")
        print(str(exc))
        print(f"\nПодробности записаны в папку: {paths['logs']}")
        wait_key()
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
